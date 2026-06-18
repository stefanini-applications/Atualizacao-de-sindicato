#!/usr/bin/env python3
"""
Gera o template de decisões em Excel (.xlsx) e CSV a partir da fila de revisão
de parâmetros sindicais (PRJ-70 / PRJ-71).

Lê `reports/parametros_revisao.json` e grava:
  - `reports/review_decisions_template.xlsx`  ← entrega operacional principal (PRJ-71)
  - `reports/review_decisions_template.csv`   ← apoio técnico (PRJ-70, mantido)

O XLSX segue o modelo de negócio enviado pelo time de Pricing/RH: uma linha por
registro/sindicato com colunas de parâmetros pré-preenchidas e colunas de revisão
editáveis ao final.

⚠️  Este script é estritamente operacional — geração de template:
    - Não aplica decisões à base de dados.
    - `data/base_parametros_sindicais.json` e `.js` não são lidos nem escritos.
    - `app.js`, `index.html` e `style.css` não são tocados.

Uso:
    python3 generate_review_decisions_template.py [--dry-run]

Opções:
    --dry-run   Exibe totais no terminal sem criar ou modificar arquivos.
"""

import argparse
import csv
import json
import os
import sys
from collections import OrderedDict

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
QUEUE_PATH = os.path.join(REPO_ROOT, "reports", "parametros_revisao.json")
REPORTS_DIR = os.path.join(REPO_ROOT, "reports")
OUTPUT_PATH = os.path.join(REPORTS_DIR, "review_decisions_template.csv")
XLSX_OUTPUT_PATH = os.path.join(REPORTS_DIR, "review_decisions_template.xlsx")

# Ordem exata das 22 colunas (AC1)
CSV_COLUMNS = [
    "registro_id",
    "uf",
    "sindicato",
    "categoria",
    "ano",
    "campo",
    "valor_atual",
    "status_atual",
    "origem",
    "fonte",
    "fonte_textual",
    "data_extracao",
    "observacao",
    "opcoes_identificadas",
    "prioridade_revisao",
    "acao_sugerida",
    "decisao_sugerida",
    "decisao_final",
    "valor_revisado",
    "observacao_revisor",
    "revisor",
    "data_revisao",
]

# Mapeamento acao_sugerida → decisao_sugerida (AC2)
DECISAO_MAP: dict[str, str] = {
    "validar": "validar",
    "revisar_conflito": "marcar_conflito",
    "buscar_fonte": "buscar_fonte",
    "manter_pendente": "manter_pendente",
}

# Sentinelas que representam ausência real de valor (AC3)
_VALOR_AUSENTE = frozenset({"Não identificado", "", None})

# ──────────────────────────────────────────────────────────────────────────────
# Constantes para o XLSX (modelo de negócio — PRJ-71)
# ──────────────────────────────────────────────────────────────────────────────

# Colunas do modelo de negócio (seguem o layout da planilha enviada pelo time)
XLSX_BUSINESS_COLS: list[str] = [
    "CODIGO DO SINDICATO",
    "ESTADO/ SINDICATO",
    "SINDICATO",
    "ESTADO",
    "APLICÁVEL Á:",
    "HR SegSex",
    "HR Sabado",
    "HR Domingo",
    "ADICIONAL NOTURNO",
    "Sobreaviso",
    "JORNADA DE TRABALHO",
    "DATA VIGENCIA PISO",
    "INÍCIO VIGÊNCIA CCT",
    "FIM VIGÊNCIA CCT/ACT",
    "PISO TECNICO SUPORTE JR",
    "Piso administrativo",
    "VR Remuneração",
    "VR Custo",
    "Salário <= 2999,99",
    "Salário >= 3000,00",
    "Seguro",
    "VT mensal",
    "OUTROS CUSTOS",
    "PLR",
    "DEFLATOR APP",
    "DEFLATOR ITO",
    "DEFLATOR BPO",
    "ATUALIZAÇÃO",
]

# Colunas operacionais de revisão (adicionadas ao final conforme PRJ-71)
XLSX_REVIEW_COLS: list[str] = [
    "status_parametro",
    "origem",
    "fonte",
    "fonte_textual",
    "opcoes_identificadas",
    "prioridade_revisao",
    "acao_sugerida",
    "decisao_sugerida",
    "decisao_final",
    "valor_revisado",
    "observacao_revisor",
    "revisor",
    "data_revisao",
]

XLSX_COLUMNS: list[str] = XLSX_BUSINESS_COLS + XLSX_REVIEW_COLS

# Mapeamento de campo JSON → coluna do modelo XLSX
CAMPO_TO_XLSX_COL: dict[str, str] = {
    "piso_salarial": "Piso administrativo",
    "adicional_noturno": "ADICIONAL NOTURNO",
    "auxilio_alimentacao": "VR Remuneração",
    "plr": "PLR",
    "hora_extra": "HR SegSex",
    "sobreaviso": "Sobreaviso",
    "jornada": "JORNADA DE TRABALHO",
}

# Prioridades para selecionar o item mais crítico de cada registro
_ACAO_PRIORITY: dict[str, int] = {
    "revisar_conflito": 4,
    "buscar_fonte": 3,
    "manter_pendente": 2,
    "validar": 1,
}
_PRIORIDADE_PRIORITY: dict[str, int] = {"alta": 3, "média": 2, "baixa": 1}
_STATUS_PRIORITY: dict[str, int] = {
    "conflito": 3,
    "pendente_revisao": 2,
    "extraido_para_revisao": 1,
}

# Cores do cabeçalho XLSX (azul escuro / texto branco)
_HEADER_BG = "1F3864"
_HEADER_FG = "FFFFFF"


# ──────────────────────────────────────────────────────────────────────────────
# Lógica de mapeamento
# ──────────────────────────────────────────────────────────────────────────────

def _map_decisao_sugerida(acao_sugerida: str | None) -> str:
    """Mapeia acao_sugerida para decisao_sugerida conforme AC2."""
    return DECISAO_MAP.get(acao_sugerida or "", "manter_pendente")


def _calc_valor_revisado(valor_atual) -> str:
    """
    Retorna o valor_atual como string quando é um valor real.
    Retorna string vazia quando é nulo, vazio ou 'Não identificado' (AC3).
    """
    if valor_atual is None or valor_atual in _VALOR_AUSENTE:
        return ""
    return str(valor_atual)


def _serialize_opcoes(opcoes) -> str:
    """
    Serializa opcoes_identificadas para string legível no CSV (AC4).
    Listas e objetos são convertidos para JSON; outros tipos ficam como string.
    """
    if opcoes is None:
        return ""
    if isinstance(opcoes, (list, dict)):
        return json.dumps(opcoes, ensure_ascii=False)
    return str(opcoes)


# ──────────────────────────────────────────────────────────────────────────────
# Construção das linhas do template
# ──────────────────────────────────────────────────────────────────────────────

def build_template_rows(itens: list[dict]) -> list[dict]:
    """
    Converte itens da fila em linhas do template CSV.
    Cada item da fila origina exatamente uma linha (AC1).
    """
    rows = []
    for item in itens:
        decisao_sugerida = _map_decisao_sugerida(item.get("acao_sugerida"))
        valor_revisado = _calc_valor_revisado(item.get("valor"))

        row = {
            "registro_id": item.get("registro_id", ""),
            "uf": item.get("uf", ""),
            "sindicato": item.get("sindicato", ""),
            "categoria": item.get("categoria", ""),
            "ano": item.get("ano", ""),
            "campo": item.get("campo", ""),
            "valor_atual": "" if item.get("valor") is None else item.get("valor"),
            "status_atual": item.get("status_parametro", ""),
            "origem": item.get("origem", ""),
            "fonte": item.get("fonte", ""),
            "fonte_textual": item.get("fonte_textual", ""),
            "data_extracao": item.get("data_extracao", ""),
            "observacao": item.get("observacao", ""),
            "opcoes_identificadas": _serialize_opcoes(item.get("opcoes_identificadas")),
            "prioridade_revisao": item.get("prioridade_revisao", ""),
            "acao_sugerida": item.get("acao_sugerida", ""),
            "decisao_sugerida": decisao_sugerida,
            "decisao_final": decisao_sugerida,  # ponto de partida editável (AC2)
            "valor_revisado": valor_revisado,
            "observacao_revisor": "",
            "revisor": "",
            "data_revisao": "",
        }
        rows.append(row)
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# Contagens para dry-run (AC5)
# ──────────────────────────────────────────────────────────────────────────────

def _count_decisoes(rows: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {
        "validar": 0,
        "marcar_conflito": 0,
        "buscar_fonte": 0,
        "manter_pendente": 0,
    }
    for row in rows:
        decisao = row.get("decisao_sugerida", "")
        if decisao in counts:
            counts[decisao] += 1
    return counts


# ──────────────────────────────────────────────────────────────────────────────
# Funções para o XLSX (PRJ-71)
# ──────────────────────────────────────────────────────────────────────────────

def _item_criticality(item: dict) -> tuple[int, int, int]:
    """Retorna tupla de prioridade para selecionar o item mais crítico do registro."""
    return (
        _STATUS_PRIORITY.get(item.get("status_parametro", ""), 0),
        _PRIORIDADE_PRIORITY.get(item.get("prioridade_revisao", ""), 0),
        _ACAO_PRIORITY.get(item.get("acao_sugerida", ""), 0),
    )


def pivot_to_registros(itens: list[dict]) -> OrderedDict:
    """Agrupa itens por registro_id preservando a ordem de aparecimento."""
    registros: OrderedDict = OrderedDict()
    for item in itens:
        rid = item.get("registro_id", "")
        registros.setdefault(rid, []).append(item)
    return registros


def build_xlsx_rows(itens: list[dict]) -> list[dict]:
    """
    Constrói linhas para o XLSX: uma linha por registro/sindicato.

    As colunas de parâmetros recebem o valor do campo correspondente.
    As colunas de revisão são derivadas do item mais crítico do registro.
    """
    registros = pivot_to_registros(itens)
    rows = []
    for rid, items in registros.items():
        ref = items[0]  # campos de identidade são iguais para o mesmo registro
        critical = max(items, key=_item_criticality)

        row: dict = {col: "" for col in XLSX_COLUMNS}

        # Colunas de identidade
        row["CODIGO DO SINDICATO"] = rid
        uf = ref.get("uf") or ""
        sindicato = ref.get("sindicato") or ""
        row["ESTADO/ SINDICATO"] = f"{uf} - {sindicato}".strip(" -")
        row["SINDICATO"] = sindicato
        row["ESTADO"] = uf
        row["APLICÁVEL Á:"] = ref.get("categoria") or ""

        # Colunas de parâmetros: mapeia cada campo ao campo Excel correto
        for item in items:
            campo = item.get("campo", "")
            xl_col = CAMPO_TO_XLSX_COL.get(campo)
            if xl_col:
                valor = item.get("valor")
                if valor not in _VALOR_AUSENTE:
                    row[xl_col] = valor

        # Colunas de revisão derivadas do item mais crítico
        decisao_sugerida = _map_decisao_sugerida(critical.get("acao_sugerida"))
        row["status_parametro"] = critical.get("status_parametro") or ""
        row["origem"] = critical.get("origem") or ""
        row["fonte"] = critical.get("fonte") or ""
        row["fonte_textual"] = critical.get("fonte_textual") or ""
        row["opcoes_identificadas"] = _serialize_opcoes(critical.get("opcoes_identificadas"))
        row["prioridade_revisao"] = critical.get("prioridade_revisao") or ""
        row["acao_sugerida"] = critical.get("acao_sugerida") or ""
        row["decisao_sugerida"] = decisao_sugerida
        row["decisao_final"] = decisao_sugerida  # ponto de partida editável
        row["valor_revisado"] = _calc_valor_revisado(critical.get("valor"))
        # observacao_revisor, revisor e data_revisao iniciam vazios para preenchimento humano

        rows.append(row)
    return rows


def save_xlsx(rows: list[dict], path: str) -> None:
    """Grava o template XLSX com cabeçalho formatado e filtros habilitados."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revisão de Parâmetros"

    # Cabeçalho
    ws.append(XLSX_COLUMNS)
    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=_HEADER_FG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    last_col_letter = get_column_letter(len(XLSX_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Linhas de dados
    for row in rows:
        ws.append([row.get(col, "") for col in XLSX_COLUMNS])

    # Larguras das colunas
    _WIDE_COLS = {"fonte_textual", "opcoes_identificadas", "APLICÁVEL Á:", "SINDICATO"}
    for col_idx, col_name in enumerate(XLSX_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in _WIDE_COLS:
            ws.column_dimensions[col_letter].width = 40
        elif col_name in ("ESTADO/ SINDICATO", "observacao_revisor"):
            ws.column_dimensions[col_letter].width = 28
        else:
            ws.column_dimensions[col_letter].width = 18

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


# ──────────────────────────────────────────────────────────────────────────────


def load_queue(path: str) -> list[dict]:
    """Lê a fila de revisão e retorna a lista de itens."""
    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)
    return data.get("itens", [])


def save_template(rows: list[dict], path: str) -> None:
    """Grava o template CSV em disco."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _print_dry_run_summary(total: int, counts: dict[str, int]) -> None:
    """Exibe o sumário obrigatório do dry-run (AC5)."""
    print(f"Total de itens lidos da fila: {total}")
    print(f"decisao_sugerida = validar: {counts['validar']}")
    print(f"decisao_sugerida = marcar_conflito: {counts['marcar_conflito']}")
    print(f"decisao_sugerida = buscar_fonte: {counts['buscar_fonte']}")
    print(f"decisao_sugerida = manter_pendente: {counts['manter_pendente']}")


# ──────────────────────────────────────────────────────────────────────────────
# Entrada principal
# ──────────────────────────────────────────────────────────────────────────────

def _parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera template CSV pré-preenchido de decisões para revisão de parâmetros sindicais (PRJ-70)."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Exibe totais no terminal sem criar ou modificar arquivos.",
    )
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = _parse_args(argv)

    if not os.path.exists(QUEUE_PATH):
        print(f"❌  Arquivo não encontrado: {QUEUE_PATH}", file=sys.stderr)
        return 1

    itens = load_queue(QUEUE_PATH)
    rows = build_template_rows(itens)
    counts = _count_decisoes(rows)

    if args.dry_run:
        print("⚠️  Modo dry-run: nenhum arquivo será criado ou modificado.\n")
        _print_dry_run_summary(len(rows), counts)
    else:
        # Entrega operacional principal: XLSX (PRJ-71)
        xlsx_rows = build_xlsx_rows(itens)
        save_xlsx(xlsx_rows, XLSX_OUTPUT_PATH)
        print(f"✅  Template Excel gravado em: {XLSX_OUTPUT_PATH}")
        print(f"    Total de registros: {len(xlsx_rows)}")

        # Apoio técnico: CSV (PRJ-70, mantido)
        save_template(rows, OUTPUT_PATH)
        print(f"✅  Template CSV gravado em:   {OUTPUT_PATH}")
        print(f"    Total de linhas: {len(rows)}")

        _print_dry_run_summary(len(rows), counts)

    return 0


if __name__ == "__main__":
    sys.exit(main())
