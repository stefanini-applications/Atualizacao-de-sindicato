#!/usr/bin/env python3
"""
Aplica decisões humanas do Excel de revisão na base sindical com auditoria completa.

Lê `reports/review_decisions_template.xlsx` (ou outro arquivo .xlsx informado via
`--decisions`) e aplica as decisões explícitas do revisor em
`data/base_parametros_sindicais.json` e `data/base_parametros_sindicais.js`,
gerando `reports/review_decisions_audit.json` com o registro de auditoria
completo (antes/depois de cada alteração).

Formatos de Excel suportados
──────────────────────────────
1. Formato por campo (recomendado): uma linha por (registro × campo).
   Colunas obrigatórias: registro_id, campo, decisao_final, valor_revisado,
   observacao_revisor, revisor, data_revisao.

2. Formato por registro (compatível com review_decisions_template.xlsx gerado por
   generate_review_decisions_template.py): uma linha por registro/sindicato.
   Colunas obrigatórias: CODIGO DO SINDICATO, decisao_final, observacao_revisor,
   revisor, data_revisao. As colunas de parâmetros (Piso administrativo, ADICIONAL
   NOTURNO, etc.) são expandidas internamente para entradas por campo usando o
   mapeamento CAMPO_TO_XLSX_COL de generate_review_decisions_template.py.

Se qualquer coluna obrigatória estiver ausente no formato detectado, o script
aborta sem alterar nenhum arquivo.

Uso:
    python3 apply_review_decisions.py --decisions reports/review_decisions_template.xlsx
    python3 apply_review_decisions.py --decisions reports/review_decisions_template.xlsx --dry-run

Opções:
    --decisions <arquivo.xlsx>  Caminho para o Excel revisado (padrão: reports/review_decisions_template.xlsx).
    --dry-run                   Exibe o sumário no terminal sem alterar ou criar
                                nenhum arquivo.

⚠️  Apenas a decisão `validar` pode setar `status_parametro: "valido"`.
    Toda alteração gera registro de auditoria com antes/depois.
    Dry-run não toca absolutamente nenhum arquivo.
"""

import argparse
import json
import os
import sys
import tempfile
from datetime import datetime, timezone

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(REPO_ROOT, "data", "base_parametros_sindicais.json")
JS_PATH = os.path.join(REPO_ROOT, "data", "base_parametros_sindicais.js")
DEFAULT_DECISIONS_PATH = os.path.join(REPO_ROOT, "reports", "review_decisions_template.xlsx")
AUDIT_PATH = os.path.join(REPO_ROOT, "reports", "review_decisions_audit.json")

# Colunas obrigatórias no formato por campo (uma linha por registro×campo)
REQUIRED_COLS = {
    "registro_id",
    "campo",
    "decisao_final",
    "valor_revisado",
    "observacao_revisor",
    "revisor",
    "data_revisao",
}

# Colunas obrigatórias no formato por registro (review_decisions_template.xlsx)
REQUIRED_COLS_XLSX = {
    "CODIGO DO SINDICATO",
    "decisao_final",
    "observacao_revisor",
    "revisor",
    "data_revisao",
}

# Mapeamento campo JSON → coluna do modelo XLSX (mesmo de generate_review_decisions_template.py)
CAMPO_TO_XLSX_COL: dict[str, str] = {
    "piso_salarial": "Piso administrativo",
    "adicional_noturno": "ADICIONAL NOTURNO",
    "auxilio_alimentacao": "VR Remuneração",
    "plr": "PLR",
    "hora_extra": "HR SegSex",
    "sobreaviso": "Sobreaviso",
    "jornada": "JORNADA DE TRABALHO",
}

# Decisões válidas e seus status resultantes
VALID_DECISIONS = {"validar", "manter_pendente", "rejeitar", "marcar_conflito", "buscar_fonte"}

# Status resultante para cada decisão
DECISION_TO_STATUS = {
    "validar": "valido",
    "manter_pendente": "pendente_revisao",
    "rejeitar": "rejeitado",
    "marcar_conflito": "conflito",
    "buscar_fonte": "pendente_revisao",
}


# ──────────────────────────────────────────────────────────────────────────────
# Leitura do Excel
# ──────────────────────────────────────────────────────────────────────────────

def _abort_missing_cols(missing: set[str]) -> None:
    missing_str = ", ".join(sorted(missing))
    print(
        f"❌  Colunas obrigatórias ausentes no Excel: {missing_str}\n"
        "    Nenhum arquivo foi alterado.",
        file=sys.stderr,
    )
    sys.exit(1)


def _expand_xlsx_rows(raw_rows: list[dict], headers_set: set[str]) -> list[dict]:
    """
    Expande linhas do formato por-registro para o formato normalizado por campo.

    Cada linha original gera uma entrada por campo presente em CAMPO_TO_XLSX_COL
    cujo valor na linha não seja nulo, ou para todos os campos se nenhum valor
    de negócio estiver presente (para permitir decisões sobre campos sem valor).
    """
    result: list[dict] = []
    for raw in raw_rows:
        registro_id = raw.get("CODIGO DO SINDICATO")
        decisao = raw.get("decisao_final")
        observacao = raw.get("observacao_revisor") or ""
        revisor = raw.get("revisor") or ""
        data_rev = raw.get("data_revisao") or ""

        for campo_name, xlsx_col in CAMPO_TO_XLSX_COL.items():
            if xlsx_col not in headers_set:
                continue
            valor_revisado = raw.get(xlsx_col)
            result.append({
                "registro_id": registro_id,
                "campo": campo_name,
                "decisao_final": decisao,
                "valor_revisado": valor_revisado,
                "observacao_revisor": observacao,
                "revisor": revisor,
                "data_revisao": data_rev,
            })

    return result


def load_decisions_xlsx(path: str) -> list[dict]:
    """
    Lê o Excel revisado e retorna uma lista normalizada de dicts por campo.

    Suporta dois formatos:
    - Por campo: colunas `registro_id` e `campo` presentes → uma linha por campo.
    - Por registro: coluna `CODIGO DO SINDICATO` presente → expande automaticamente
      para entradas por campo usando CAMPO_TO_XLSX_COL.

    Aborta com SystemExit(1) se qualquer coluna obrigatória estiver ausente.
    """
    try:
        import openpyxl
    except ImportError:
        print("❌  Dependência ausente: openpyxl. Instale com: apt install python3-openpyxl", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(path):
        print(f"❌  Arquivo não encontrado: {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    headers_set = {h for h in headers if h is not None}
    col_index = {name: idx for idx, name in enumerate(headers) if name is not None}

    # Detecta formato
    is_per_campo = ("registro_id" in headers_set and "campo" in headers_set)

    if is_per_campo:
        missing = REQUIRED_COLS - headers_set
        if missing:
            _abort_missing_cols(missing)
    elif "CODIGO DO SINDICATO" in headers_set:
        missing = REQUIRED_COLS_XLSX - headers_set
        if missing:
            _abort_missing_cols(missing)
    else:
        print(
            "❌  Formato de Excel não reconhecido.\n"
            "    O arquivo deve conter 'registro_id' + 'campo' (formato por campo)\n"
            "    ou 'CODIGO DO SINDICATO' (formato por registro).\n"
            "    Nenhum arquivo foi alterado.",
            file=sys.stderr,
        )
        sys.exit(1)

    raw_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        entry = {col: row[idx] for col, idx in col_index.items()}
        raw_rows.append(entry)

    if is_per_campo:
        return raw_rows
    return _expand_xlsx_rows(raw_rows, headers_set)


# ──────────────────────────────────────────────────────────────────────────────
# Leitura e escrita da base JSON
# ──────────────────────────────────────────────────────────────────────────────

def load_base(path: str) -> dict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def _save_json_atomic(data: dict, path: str) -> None:
    """Salva JSON de forma atômica: escreve em temp e renomeia."""
    dir_ = os.path.dirname(path)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=dir_, delete=False, suffix=".tmp") as tmp:
        json.dump(data, tmp, ensure_ascii=False, indent=2)
        tmp_path = tmp.name
    os.replace(tmp_path, path)


def _save_js(data: dict, path: str) -> None:
    """Regenera o arquivo .js a partir do JSON (equivalente a export_inline_data.py)."""
    js_content = (
        "// Gerado automaticamente por apply_review_decisions.py — não editar manualmente.\n"
        "window.BASE_PARAMETROS_SINDICAIS = "
        + json.dumps(data, ensure_ascii=False)
        + ";\n"
    )
    dir_ = os.path.dirname(path)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=dir_, delete=False, suffix=".tmp") as tmp:
        tmp.write(js_content)
        tmp_path = tmp.name
    os.replace(tmp_path, path)


# ──────────────────────────────────────────────────────────────────────────────
# Motor de aplicação de decisões
# ──────────────────────────────────────────────────────────────────────────────

def _build_registro_index(base: dict) -> dict[str, dict]:
    """Retorna dict {id_registro_reajuste: registro} para busca O(1)."""
    return {reg["id_registro_reajuste"]: reg for reg in base.get("registros", [])}


def _normalize_valor(v) -> str | None:
    """Normaliza valor para comparação: None/string vazia → None."""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    return str(v)


def _coerce_valor(new_val, current_val):
    """
    Tenta converter new_val para o mesmo tipo de current_val.
    Se falhar, retorna new_val como está.
    """
    if new_val is None:
        return None
    if current_val is not None and not isinstance(current_val, str):
        try:
            return type(current_val)(new_val)
        except (ValueError, TypeError):
            pass
    return new_val


def apply_decision_to_campo(
    campo_data: dict,
    decisao: str,
    valor_revisado,
    revisor: str,
    data_revisao: str,
    observacao_revisor: str,
    timestamp: str,
) -> tuple[dict, dict]:
    """
    Aplica uma decisão a um campo da base e retorna (campo_atualizado, audit_delta).

    audit_delta contém os campos que mudaram (valor_anterior / valor_novo,
    status_anterior / status_novo).
    """
    status_anterior = campo_data.get("status_parametro")
    valor_anterior = campo_data.get("valor")

    novo_campo = dict(campo_data)
    status_novo = DECISION_TO_STATUS[decisao]
    valor_novo = valor_anterior  # por padrão, valor não muda

    if decisao == "validar":
        novo_campo["status_parametro"] = "valido"
        novo_campo["validado_por"] = revisor
        novo_campo["data_validacao"] = data_revisao
        novo_campo["observacao_validacao"] = observacao_revisor or None

        valor_norm = _normalize_valor(valor_revisado)
        if valor_norm is not None:
            coerced = _coerce_valor(valor_revisado, valor_anterior)
            # Só atualiza se diferente do valor atual
            if coerced != valor_anterior:
                novo_campo["valor_original_pre_validacao"] = valor_anterior
                novo_campo["valor"] = coerced
                valor_novo = coerced

    elif decisao == "manter_pendente":
        novo_campo["status_parametro"] = "pendente_revisao"

    elif decisao == "rejeitar":
        novo_campo["status_parametro"] = "rejeitado"

    elif decisao == "marcar_conflito":
        novo_campo["status_parametro"] = "conflito"
        # Preserva opcoes_identificadas existentes — não sobrescrever

    elif decisao == "buscar_fonte":
        novo_campo["status_parametro"] = "pendente_revisao"
        novo_campo["acao_recomendada"] = "buscar_fonte"

    audit_delta = {
        "status_anterior": status_anterior,
        "status_novo": status_novo,
        "valor_anterior": valor_anterior,
        "valor_novo": valor_novo,
        "timestamp_execucao": timestamp,
    }

    return novo_campo, audit_delta


def process_decisions(
    rows: list[dict],
    base: dict,
    timestamp: str,
) -> tuple[dict, list[dict], dict[str, int]]:
    """
    Processa todas as linhas do Excel contra a base.

    Retorna:
        base_modificada  — cópia da base com decisões válidas aplicadas
        audit_records    — lista de registros de auditoria (um por linha)
        summary          — contagens por categoria
    """
    import copy

    base_mod = copy.deepcopy(base)
    reg_index = _build_registro_index(base_mod)

    audit_records: list[dict] = []
    summary: dict[str, int] = {
        "total_lidas": 0,
        "validar": 0,
        "manter_pendente": 0,
        "rejeitar": 0,
        "marcar_conflito": 0,
        "buscar_fonte": 0,
        "ignoradas": 0,
        "erros": 0,
    }

    for row in rows:
        summary["total_lidas"] += 1

        registro_id = row.get("registro_id")
        campo = row.get("campo")
        decisao = row.get("decisao_final")
        valor_revisado = row.get("valor_revisado")
        observacao_revisor = row.get("observacao_revisor") or ""
        revisor = row.get("revisor") or ""
        data_revisao = str(row.get("data_revisao") or "")

        base_audit: dict = {
            "registro_id": registro_id,
            "campo": campo,
            "decisao_final": decisao,
            "revisor": revisor,
            "data_revisao": data_revisao,
            "observacao_revisor": observacao_revisor,
            "timestamp_execucao": timestamp,
        }

        # Valida decisão
        if not decisao or str(decisao).strip() == "":
            summary["ignoradas"] += 1
            audit_records.append({**base_audit, "resultado": "ignorado", "motivo": "decisao_final vazia"})
            continue

        decisao_str = str(decisao).strip().lower()

        if decisao_str not in VALID_DECISIONS:
            summary["erros"] += 1
            audit_records.append({
                **base_audit,
                "resultado": "erro",
                "motivo": f"decisao_final inválida: '{decisao}'",
                "status_anterior": None,
                "status_novo": None,
                "valor_anterior": None,
                "valor_novo": None,
            })
            continue

        # Valida registro_id
        if not registro_id or str(registro_id).strip() == "":
            summary["erros"] += 1
            audit_records.append({
                **base_audit,
                "resultado": "erro",
                "motivo": "registro_id vazio",
                "status_anterior": None,
                "status_novo": None,
                "valor_anterior": None,
                "valor_novo": None,
            })
            continue

        reg = reg_index.get(str(registro_id).strip())
        if reg is None:
            summary["erros"] += 1
            audit_records.append({
                **base_audit,
                "resultado": "erro",
                "motivo": f"registro_id não encontrado na base: '{registro_id}'",
                "status_anterior": None,
                "status_novo": None,
                "valor_anterior": None,
                "valor_novo": None,
            })
            continue

        # Valida campo
        if not campo or str(campo).strip() == "":
            summary["erros"] += 1
            audit_records.append({
                **base_audit,
                "resultado": "erro",
                "motivo": "campo vazio",
                "status_anterior": None,
                "status_novo": None,
                "valor_anterior": None,
                "valor_novo": None,
            })
            continue

        campo_str = str(campo).strip()
        itens_cct = reg.get("itens_cct", {})
        if campo_str not in itens_cct:
            summary["erros"] += 1
            audit_records.append({
                **base_audit,
                "resultado": "erro",
                "motivo": f"campo '{campo_str}' não encontrado em itens_cct do registro '{registro_id}'",
                "status_anterior": None,
                "status_novo": None,
                "valor_anterior": None,
                "valor_novo": None,
            })
            continue

        # Aplica decisão
        campo_atual = itens_cct[campo_str]
        novo_campo, delta = apply_decision_to_campo(
            campo_data=campo_atual,
            decisao=decisao_str,
            valor_revisado=valor_revisado,
            revisor=revisor,
            data_revisao=data_revisao,
            observacao_revisor=observacao_revisor,
            timestamp=timestamp,
        )

        reg["itens_cct"][campo_str] = novo_campo
        summary[decisao_str] += 1

        audit_records.append({
            **base_audit,
            "resultado": "aplicado",
            "motivo": None,
            **delta,
        })

    return base_mod, audit_records, summary


# ──────────────────────────────────────────────────────────────────────────────
# Persistência da auditoria
# ──────────────────────────────────────────────────────────────────────────────

def build_audit_report(audit_records: list[dict], summary: dict, timestamp: str) -> dict:
    return {
        "timestamp_execucao": timestamp,
        "total_lidas": summary["total_lidas"],
        "resumo": {
            "validar": summary["validar"],
            "manter_pendente": summary["manter_pendente"],
            "rejeitar": summary["rejeitar"],
            "marcar_conflito": summary["marcar_conflito"],
            "buscar_fonte": summary["buscar_fonte"],
            "ignoradas": summary["ignoradas"],
            "erros": summary["erros"],
        },
        "registros": audit_records,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Sumário dry-run
# ──────────────────────────────────────────────────────────────────────────────

def _print_summary(summary: dict, dry_run: bool) -> None:
    prefix = "⚠️  [DRY-RUN] " if dry_run else "✅  "
    print(f"{prefix}Decisões lidas:              {summary['total_lidas']}")
    print(f"    validar:                     {summary['validar']}")
    print(f"    manter_pendente:             {summary['manter_pendente']}")
    print(f"    rejeitar:                    {summary['rejeitar']}")
    print(f"    marcar_conflito:             {summary['marcar_conflito']}")
    print(f"    buscar_fonte:                {summary['buscar_fonte']}")
    print(f"    ignoradas:                   {summary['ignoradas']}")
    print(f"    erros:                       {summary['erros']}")


# ──────────────────────────────────────────────────────────────────────────────
# Entrada principal
# ──────────────────────────────────────────────────────────────────────────────

def _parse_args(argv=None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Aplica decisões humanas do Excel de revisão na base sindical "
            "com auditoria completa (PRJ-72)."
        )
    )
    parser.add_argument(
        "--decisions",
        default=DEFAULT_DECISIONS_PATH,
        metavar="<arquivo.xlsx>",
        help="Caminho para o Excel revisado (padrão: reports/review_decisions_template.xlsx).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Exibe o sumário no terminal sem alterar ou criar nenhum arquivo.",
    )
    return parser.parse_args(argv)


def main(argv=None) -> int:
    args = _parse_args(argv)

    # 1. Lê e valida o Excel
    rows = load_decisions_xlsx(args.decisions)

    # 2. Lê a base atual
    if not os.path.exists(JSON_PATH):
        print(f"❌  Base não encontrada: {JSON_PATH}", file=sys.stderr)
        return 1
    base = load_base(JSON_PATH)

    timestamp = datetime.now(timezone.utc).isoformat()

    # 3. Processa decisões em memória (não toca disco)
    base_mod, audit_records, summary = process_decisions(rows, base, timestamp)

    # 4. Dry-run: apenas exibe sumário
    if args.dry_run:
        print("⚠️  Modo dry-run: nenhum arquivo será criado ou modificado.\n")
        _print_summary(summary, dry_run=True)
        return 0

    # 5. Execução real — ordem obrigatória de escrita:
    #    (a) auditoria construída em memória ✓ (já feito acima)
    #    (b) base_parametros_sindicais.json
    _save_json_atomic(base_mod, JSON_PATH)
    print(f"✅  Base JSON atualizada: {JSON_PATH}")

    #    (c) base_parametros_sindicais.js
    _save_js(base_mod, JS_PATH)
    print(f"✅  Base JS regenerada:   {JS_PATH}")

    #    (d) reports/review_decisions_audit.json
    audit_report = build_audit_report(audit_records, summary, timestamp)
    os.makedirs(os.path.dirname(AUDIT_PATH), exist_ok=True)
    _save_json_atomic(audit_report, AUDIT_PATH)
    print(f"✅  Auditoria gravada:    {AUDIT_PATH}")

    _print_summary(summary, dry_run=False)

    return 0


if __name__ == "__main__":
    sys.exit(main())
