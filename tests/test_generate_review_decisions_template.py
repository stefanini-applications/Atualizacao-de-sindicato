"""
Testes automatizados para generate_review_decisions_template.py (PRJ-70).

Cobre os 10 cenários obrigatórios definidos nos ACs:
  1.  Template gerado a partir de parametros_revisao.json (AC1)
  2.  decisao_sugerida correta para cada uma das 4 ações (AC2)
  3.  decisao_final inicia igual a decisao_sugerida (AC2)
  4.  valor_revisado preenchido quando valor_atual existe e é valor real (AC3)
  5.  valor_revisado vazio quando valor_atual é nulo (AC3)
  6.  valor_revisado vazio quando valor_atual é string vazia (AC3)
  7.  valor_revisado vazio quando valor_atual é "Não identificado" (AC3)
  8.  fonte_textual, origem, observacao e opcoes_identificadas preservados (AC4)
  9.  --dry-run não cria arquivo e exibe totais corretos (AC5)
  10. base_parametros_sindicais.json e .js não são alterados (AC6)
"""

import copy
import csv
import json
import os
import sys
import tempfile
import unittest
from io import StringIO
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from generate_review_decisions_template import (
    BUSINESS_COLUMNS,
    CSV_COLUMNS,
    DECISAO_MAP,
    REVIEW_COLUMNS,
    XLSX_COLUMNS,
    _calc_valor_revisado,
    _count_decisoes,
    _map_decisao_sugerida,
    _serialize_opcoes,
    build_template_rows,
    load_queue,
    main,
    save_template,
    save_xlsx,
)

# ──────────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────────

ITEM_VALIDAR = {
    "registro_id": "REG-SP-TEST-2025",
    "uf": "SP",
    "sindicato": "Sindicato Teste SP",
    "categoria": "Tecnologia",
    "ano": 2025,
    "campo": "piso_salarial",
    "valor": 1540.47,
    "status_parametro": "extraido_para_revisao",
    "origem": "pdf_cct",
    "fonte": "PDF da CCT",
    "fonte_textual": "CLÁUSULA TERCEIRA - PISO SALARIAL R$ 1.540,47",
    "data_extracao": "2026-06-15",
    "observacao": None,
    "opcoes_identificadas": None,
    "prioridade_revisao": "média",
    "acao_sugerida": "validar",
}

ITEM_REVISAR_CONFLITO = {
    **ITEM_VALIDAR,
    "campo": "hora_extra",
    "valor": 1600.0,
    "status_parametro": "conflito",
    "origem": "conflito_pdf_mte",
    "observacao": "Divergência PDF vs MTE",
    "opcoes_identificadas": [1540.47, 1620.0],
    "prioridade_revisao": "alta",
    "acao_sugerida": "revisar_conflito",
}

ITEM_BUSCAR_FONTE = {
    **ITEM_VALIDAR,
    "campo": "vr",
    "valor": None,
    "status_parametro": "pendente_revisao",
    "origem": "nao_identificado_pdf",
    "fonte": None,
    "fonte_textual": None,
    "observacao": "Cláusula não localizada",
    "opcoes_identificadas": None,
    "prioridade_revisao": "alta",
    "acao_sugerida": "buscar_fonte",
}

ITEM_MANTER_PENDENTE = {
    **ITEM_VALIDAR,
    "campo": "sobreaviso",
    "valor": None,
    "status_parametro": "extraido_para_revisao",
    "origem": "pdf_cct",
    "fonte": None,
    "fonte_textual": None,
    "observacao": None,
    "opcoes_identificadas": None,
    "prioridade_revisao": "baixa",
    "acao_sugerida": "manter_pendente",
}

ITEM_VALOR_NAO_IDENTIFICADO = {
    **ITEM_VALIDAR,
    "campo": "va",
    "valor": "Não identificado",
    "acao_sugerida": "buscar_fonte",
}

ITEM_VALOR_VAZIO = {
    **ITEM_VALIDAR,
    "campo": "plr",
    "valor": "",
    "acao_sugerida": "manter_pendente",
}

ALL_ITEMS = [
    ITEM_VALIDAR,
    ITEM_REVISAR_CONFLITO,
    ITEM_BUSCAR_FONTE,
    ITEM_MANTER_PENDENTE,
]


def _make_queue_json(itens: list[dict]) -> dict:
    return {
        "data_execucao": "2026-06-17T00:00:00+00:00",
        "dry_run": False,
        "total_itens_revisao": len(itens),
        "itens": itens,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 1 — Template gerado com todas as linhas e 22 colunas (AC1)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario1TemplateGerado(unittest.TestCase):
    def test_numero_de_linhas_igual_ao_total_da_fila(self):
        rows = build_template_rows(ALL_ITEMS)
        self.assertEqual(len(rows), len(ALL_ITEMS))

    def test_csv_tem_22_colunas(self):
        self.assertEqual(len(CSV_COLUMNS), 22)

    def test_csv_contem_todas_colunas_obrigatorias(self):
        expected = [
            "registro_id", "uf", "sindicato", "categoria", "ano", "campo",
            "valor_atual", "status_atual", "origem", "fonte", "fonte_textual",
            "data_extracao", "observacao", "opcoes_identificadas",
            "prioridade_revisao", "acao_sugerida", "decisao_sugerida",
            "decisao_final", "valor_revisado", "observacao_revisor",
            "revisor", "data_revisao",
        ]
        self.assertEqual(CSV_COLUMNS, expected)

    def test_csv_escrito_e_legivel(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8"
        ) as fh:
            path = fh.name
        try:
            save_template(rows, path)
            with open(path, encoding="utf-8", newline="") as fh:
                reader = csv.DictReader(fh)
                file_rows = list(reader)
            self.assertEqual(len(file_rows), 1)
            self.assertEqual(set(file_rows[0].keys()), set(CSV_COLUMNS))
        finally:
            os.unlink(path)

    def test_load_queue_retorna_itens(self):
        queue_data = _make_queue_json(ALL_ITEMS)
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as fh:
            json.dump(queue_data, fh, ensure_ascii=False)
            path = fh.name
        try:
            itens = load_queue(path)
            self.assertEqual(len(itens), len(ALL_ITEMS))
        finally:
            os.unlink(path)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 2 — decisao_sugerida correta para cada acao_sugerida (AC2)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario2DecisaoSugerida(unittest.TestCase):
    def test_validar_mapeia_para_validar(self):
        self.assertEqual(_map_decisao_sugerida("validar"), "validar")

    def test_revisar_conflito_mapeia_para_marcar_conflito(self):
        self.assertEqual(_map_decisao_sugerida("revisar_conflito"), "marcar_conflito")

    def test_buscar_fonte_mapeia_para_buscar_fonte(self):
        self.assertEqual(_map_decisao_sugerida("buscar_fonte"), "buscar_fonte")

    def test_manter_pendente_mapeia_para_manter_pendente(self):
        self.assertEqual(_map_decisao_sugerida("manter_pendente"), "manter_pendente")

    def test_decisao_sugerida_na_row_validar(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        self.assertEqual(rows[0]["decisao_sugerida"], "validar")

    def test_decisao_sugerida_na_row_marcar_conflito(self):
        rows = build_template_rows([copy.deepcopy(ITEM_REVISAR_CONFLITO)])
        self.assertEqual(rows[0]["decisao_sugerida"], "marcar_conflito")

    def test_decisao_sugerida_na_row_buscar_fonte(self):
        rows = build_template_rows([copy.deepcopy(ITEM_BUSCAR_FONTE)])
        self.assertEqual(rows[0]["decisao_sugerida"], "buscar_fonte")

    def test_decisao_sugerida_na_row_manter_pendente(self):
        rows = build_template_rows([copy.deepcopy(ITEM_MANTER_PENDENTE)])
        self.assertEqual(rows[0]["decisao_sugerida"], "manter_pendente")

    def test_acao_desconhecida_resulta_em_manter_pendente(self):
        self.assertEqual(_map_decisao_sugerida("acao_inexistente"), "manter_pendente")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 3 — decisao_final inicia igual a decisao_sugerida (AC2)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario3DecisaoFinal(unittest.TestCase):
    def _assert_decisao_final_igual_sugerida(self, item: dict):
        rows = build_template_rows([copy.deepcopy(item)])
        self.assertEqual(rows[0]["decisao_final"], rows[0]["decisao_sugerida"])

    def test_decisao_final_igual_sugerida_para_validar(self):
        self._assert_decisao_final_igual_sugerida(ITEM_VALIDAR)

    def test_decisao_final_igual_sugerida_para_conflito(self):
        self._assert_decisao_final_igual_sugerida(ITEM_REVISAR_CONFLITO)

    def test_decisao_final_igual_sugerida_para_buscar_fonte(self):
        self._assert_decisao_final_igual_sugerida(ITEM_BUSCAR_FONTE)

    def test_decisao_final_igual_sugerida_para_manter_pendente(self):
        self._assert_decisao_final_igual_sugerida(ITEM_MANTER_PENDENTE)

    def test_campos_editaveis_iniciam_vazios(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        row = rows[0]
        self.assertEqual(row["observacao_revisor"], "")
        self.assertEqual(row["revisor"], "")
        self.assertEqual(row["data_revisao"], "")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 4 — valor_revisado preenchido com valor real (AC3)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario4ValorRevisadoPreenchido(unittest.TestCase):
    def test_valor_numerico_preenche_valor_revisado(self):
        self.assertEqual(_calc_valor_revisado(1540.47), "1540.47")

    def test_valor_inteiro_preenche_valor_revisado(self):
        self.assertEqual(_calc_valor_revisado(1500), "1500")

    def test_valor_string_real_preenche_valor_revisado(self):
        self.assertEqual(_calc_valor_revisado("50%"), "50%")

    def test_valor_revisado_na_row_com_valor_real(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        self.assertEqual(rows[0]["valor_revisado"], "1540.47")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 5 — valor_revisado vazio quando valor é nulo (AC3)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario5ValorRevisadoNulo(unittest.TestCase):
    def test_none_resulta_em_string_vazia(self):
        self.assertEqual(_calc_valor_revisado(None), "")

    def test_valor_revisado_na_row_com_valor_nulo(self):
        rows = build_template_rows([copy.deepcopy(ITEM_BUSCAR_FONTE)])
        self.assertEqual(rows[0]["valor_revisado"], "")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 6 — valor_revisado vazio quando valor é string vazia (AC3)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario6ValorRevisadoVazio(unittest.TestCase):
    def test_string_vazia_resulta_em_string_vazia(self):
        self.assertEqual(_calc_valor_revisado(""), "")

    def test_valor_revisado_na_row_com_valor_vazio(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALOR_VAZIO)])
        self.assertEqual(rows[0]["valor_revisado"], "")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 7 — valor_revisado vazio quando valor é "Não identificado" (AC3)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario7ValorNaoIdentificado(unittest.TestCase):
    def test_nao_identificado_resulta_em_string_vazia(self):
        self.assertEqual(_calc_valor_revisado("Não identificado"), "")

    def test_valor_revisado_na_row_com_nao_identificado(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALOR_NAO_IDENTIFICADO)])
        self.assertEqual(rows[0]["valor_revisado"], "")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 8 — Campos de evidência preservados no CSV (AC4)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario8CamposPreservados(unittest.TestCase):
    def test_fonte_textual_preservada(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        self.assertEqual(rows[0]["fonte_textual"], ITEM_VALIDAR["fonte_textual"])

    def test_origem_preservada(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        self.assertEqual(rows[0]["origem"], ITEM_VALIDAR["origem"])

    def test_observacao_preservada(self):
        rows = build_template_rows([copy.deepcopy(ITEM_REVISAR_CONFLITO)])
        self.assertEqual(rows[0]["observacao"], ITEM_REVISAR_CONFLITO["observacao"])

    def test_opcoes_identificadas_lista_serializada_como_json(self):
        rows = build_template_rows([copy.deepcopy(ITEM_REVISAR_CONFLITO)])
        serialized = rows[0]["opcoes_identificadas"]
        parsed = json.loads(serialized)
        self.assertEqual(parsed, ITEM_REVISAR_CONFLITO["opcoes_identificadas"])

    def test_opcoes_identificadas_none_resulta_em_vazio(self):
        self.assertEqual(_serialize_opcoes(None), "")

    def test_opcoes_identificadas_dict_serializado(self):
        opcoes = {"pdf": 1540.47, "mte": 1620.0}
        serialized = _serialize_opcoes(opcoes)
        parsed = json.loads(serialized)
        self.assertEqual(parsed, opcoes)

    def test_opcoes_identificadas_string_preservada(self):
        self.assertEqual(_serialize_opcoes("texto simples"), "texto simples")

    def test_status_atual_mapeado_de_status_parametro(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        self.assertEqual(rows[0]["status_atual"], ITEM_VALIDAR["status_parametro"])

    def test_valor_atual_mapeado_de_valor(self):
        rows = build_template_rows([copy.deepcopy(ITEM_VALIDAR)])
        self.assertEqual(rows[0]["valor_atual"], ITEM_VALIDAR["valor"])


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 9 — --dry-run não cria arquivo e exibe totais (AC5)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario9DryRun(unittest.TestCase):
    def _run_main_dry_run(self, queue_json: dict) -> str:
        """Executa main --dry-run a partir de um JSON temporário e retorna stdout."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as fh:
            json.dump(queue_json, fh, ensure_ascii=False)
            queue_path = fh.name

        output_dir = tempfile.mkdtemp()
        output_path = os.path.join(output_dir, "review_decisions_template.csv")
        output_xlsx_path = os.path.join(output_dir, "review_decisions_template.xlsx")

        captured = StringIO()
        try:
            with (
                patch(
                    "generate_review_decisions_template.QUEUE_PATH", queue_path
                ),
                patch(
                    "generate_review_decisions_template.OUTPUT_PATH", output_path
                ),
                patch(
                    "generate_review_decisions_template.OUTPUT_XLSX_PATH",
                    output_xlsx_path,
                ),
                patch("sys.stdout", captured),
            ):
                result = main(["--dry-run"])
        finally:
            os.unlink(queue_path)

        return output_xlsx_path, captured.getvalue(), result

    def test_dry_run_nao_cria_arquivo(self):
        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        output_xlsx_path, _, exit_code = self._run_main_dry_run(queue_json)
        self.assertFalse(
            os.path.exists(output_xlsx_path),
            "dry-run não deve criar o arquivo Excel",
        )
        self.assertEqual(exit_code, 0)

    def test_dry_run_exibe_total_de_itens(self):
        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        _, output, _ = self._run_main_dry_run(queue_json)
        self.assertIn(f"Total de itens lidos da fila: {len(ALL_ITEMS)}", output)

    def test_dry_run_exibe_contagens_por_decisao(self):
        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        _, output, _ = self._run_main_dry_run(queue_json)
        self.assertIn("decisao_sugerida = validar:", output)
        self.assertIn("decisao_sugerida = marcar_conflito:", output)
        self.assertIn("decisao_sugerida = buscar_fonte:", output)
        self.assertIn("decisao_sugerida = manter_pendente:", output)

    def test_count_decisoes_totais_corretos(self):
        rows = build_template_rows(copy.deepcopy(ALL_ITEMS))
        counts = _count_decisoes(rows)
        self.assertEqual(counts["validar"], 1)
        self.assertEqual(counts["marcar_conflito"], 1)
        self.assertEqual(counts["buscar_fonte"], 1)
        self.assertEqual(counts["manter_pendente"], 1)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 10 — base_parametros_sindicais não é alterada (AC6)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario10BaseNaoAlterada(unittest.TestCase):
    def _run_main_normal(self, queue_json: dict) -> str:
        """Executa main normal a partir de um JSON temporário. Retorna path do CSV."""
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as fh:
            json.dump(queue_json, fh, ensure_ascii=False)
            queue_path = fh.name

        output_dir = tempfile.mkdtemp()
        output_path = os.path.join(output_dir, "review_decisions_template.csv")
        output_xlsx_path = os.path.join(output_dir, "review_decisions_template.xlsx")

        with (
            patch("generate_review_decisions_template.QUEUE_PATH", queue_path),
            patch("generate_review_decisions_template.OUTPUT_PATH", output_path),
            patch(
                "generate_review_decisions_template.OUTPUT_XLSX_PATH",
                output_xlsx_path,
            ),
            patch("sys.stdout", StringIO()),
        ):
            main([])

        os.unlink(queue_path)
        return output_path

    def test_base_json_nao_e_modificada(self):
        repo_root = os.path.join(os.path.dirname(__file__), "..")
        base_json = os.path.join(repo_root, "data", "base_parametros_sindicais.json")
        if not os.path.exists(base_json):
            self.skipTest("base_parametros_sindicais.json não encontrado")

        stat_before = os.stat(base_json)
        with open(base_json, "rb") as fh:
            content_before = fh.read()

        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        self._run_main_normal(queue_json)

        stat_after = os.stat(base_json)
        with open(base_json, "rb") as fh:
            content_after = fh.read()

        self.assertEqual(
            content_before,
            content_after,
            "base_parametros_sindicais.json foi modificado",
        )
        self.assertEqual(
            stat_before.st_mtime,
            stat_after.st_mtime,
            "mtime de base_parametros_sindicais.json foi alterado",
        )

    def test_base_js_nao_e_modificada(self):
        repo_root = os.path.join(os.path.dirname(__file__), "..")
        base_js = os.path.join(repo_root, "data", "base_parametros_sindicais.js")
        if not os.path.exists(base_js):
            self.skipTest("base_parametros_sindicais.js não encontrado")

        with open(base_js, "rb") as fh:
            content_before = fh.read()

        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        self._run_main_normal(queue_json)

        with open(base_js, "rb") as fh:
            content_after = fh.read()

        self.assertEqual(
            content_before,
            content_after,
            "base_parametros_sindicais.js foi modificado",
        )

    def test_script_nao_abre_arquivos_protegidos(self):
        """Garante que o script não abre nem lê arquivos protegidos em código executável."""
        import ast

        script_path = os.path.join(
            os.path.dirname(__file__), "..", "generate_review_decisions_template.py"
        )
        with open(script_path, encoding="utf-8") as fh:
            source = fh.read()

        # Remove docstrings antes de procurar por referências a arquivos protegidos
        # para não falhar em comentários de documentação legítimos.
        tree = ast.parse(source)
        non_docstring_lines = set(range(1, source.count("\n") + 2))
        for node in ast.walk(tree):
            if isinstance(node, ast.Expr) and isinstance(node.value, ast.Constant):
                for lineno in range(node.lineno, node.end_lineno + 1):
                    non_docstring_lines.discard(lineno)

        code_lines = [
            line
            for i, line in enumerate(source.splitlines(), start=1)
            if i in non_docstring_lines and not line.strip().startswith("#")
        ]
        code_only = "\n".join(code_lines)

        for protected in ["app.js", "index.html", "style.css"]:
            self.assertNotIn(
                protected,
                code_only,
                f"O script referencia o arquivo protegido '{protected}' em código executável",
            )


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 11 — Geração do arquivo .xlsx (PRJ-71)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario11XlsxGerado(unittest.TestCase):
    def _save_xlsx_to_tmp(self, itens: list[dict]) -> str:
        rows = build_template_rows(copy.deepcopy(itens))
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        save_xlsx(rows, path)
        return path

    def test_xlsx_criado_em_disco(self):
        path = self._save_xlsx_to_tmp([copy.deepcopy(ITEM_VALIDAR)])
        try:
            self.assertTrue(os.path.exists(path))
            self.assertGreater(os.path.getsize(path), 0)
        finally:
            os.unlink(path)

    def test_xlsx_e_legivel_pelo_openpyxl(self):
        from openpyxl import load_workbook

        path = self._save_xlsx_to_tmp([copy.deepcopy(ITEM_VALIDAR)])
        try:
            wb = load_workbook(path)
            self.assertIn("Revisão Sindicatos", wb.sheetnames)
        finally:
            os.unlink(path)

    def test_xlsx_contem_colunas_do_modelo_de_negocio(self):
        from openpyxl import load_workbook

        path = self._save_xlsx_to_tmp([copy.deepcopy(ITEM_VALIDAR)])
        try:
            ws = load_workbook(path).active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for expected_col in BUSINESS_COLUMNS:
                self.assertIn(expected_col, headers, f"Coluna '{expected_col}' ausente")
        finally:
            os.unlink(path)

    def test_xlsx_contem_colunas_de_revisao(self):
        from openpyxl import load_workbook

        path = self._save_xlsx_to_tmp([copy.deepcopy(ITEM_VALIDAR)])
        try:
            ws = load_workbook(path).active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for expected_col in REVIEW_COLUMNS:
                self.assertIn(expected_col, headers, f"Coluna de revisão '{expected_col}' ausente")
        finally:
            os.unlink(path)

    def test_xlsx_numero_de_colunas_correto(self):
        from openpyxl import load_workbook

        path = self._save_xlsx_to_tmp([copy.deepcopy(ITEM_VALIDAR)])
        try:
            ws = load_workbook(path).active
            self.assertEqual(ws.max_column, len(XLSX_COLUMNS))
        finally:
            os.unlink(path)

    def test_xlsx_numero_de_linhas_igual_ao_total_mais_cabecalho(self):
        from openpyxl import load_workbook

        path = self._save_xlsx_to_tmp(ALL_ITEMS)
        try:
            ws = load_workbook(path).active
            # 1 cabeçalho + len(ALL_ITEMS) linhas de dados
            self.assertEqual(ws.max_row, len(ALL_ITEMS) + 1)
        finally:
            os.unlink(path)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 12 — Preenchimento correto de decisões no Excel (PRJ-71)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario12XlsxDecisoes(unittest.TestCase):
    def _get_row_dict(self, item: dict) -> dict:
        from openpyxl import load_workbook

        rows = build_template_rows([copy.deepcopy(item)])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        save_xlsx(rows, path)
        try:
            ws = load_workbook(path).active
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            values = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        finally:
            os.unlink(path)
        return dict(zip(headers, values))

    def test_decisao_sugerida_no_xlsx(self):
        row = self._get_row_dict(ITEM_VALIDAR)
        self.assertEqual(row["decisao_sugerida"], "validar")

    def test_decisao_final_igual_sugerida_no_xlsx(self):
        row = self._get_row_dict(ITEM_VALIDAR)
        self.assertEqual(row["decisao_final"], row["decisao_sugerida"])

    def test_valor_revisado_preenchido_no_xlsx(self):
        row = self._get_row_dict(ITEM_VALIDAR)
        self.assertEqual(str(row["valor_revisado"]), "1540.47")

    def test_valor_revisado_vazio_no_xlsx_para_nulo(self):
        row = self._get_row_dict(ITEM_BUSCAR_FONTE)
        self.assertIn(row["valor_revisado"], (None, ""))

    def test_fonte_textual_preservada_no_xlsx(self):
        row = self._get_row_dict(ITEM_VALIDAR)
        self.assertEqual(row["fonte_textual"], ITEM_VALIDAR["fonte_textual"])

    def test_origem_preservada_no_xlsx(self):
        row = self._get_row_dict(ITEM_VALIDAR)
        self.assertEqual(row["origem"], ITEM_VALIDAR["origem"])

    def test_campos_editaveis_vazios_no_xlsx(self):
        row = self._get_row_dict(ITEM_VALIDAR)
        self.assertIn(row["observacao_revisor"], (None, ""))
        self.assertIn(row["revisor"], (None, ""))
        self.assertIn(row["data_revisao"], (None, ""))


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 13 — dry-run não cria arquivo .xlsx (PRJ-71)
# ──────────────────────────────────────────────────────────────────────────────

class TestCenario13XlsxDryRun(unittest.TestCase):
    def test_dry_run_nao_cria_xlsx(self):
        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as fh:
            json.dump(queue_json, fh, ensure_ascii=False)
            queue_path = fh.name

        output_dir = tempfile.mkdtemp()
        xlsx_path = os.path.join(output_dir, "review_decisions_template.xlsx")
        csv_path = os.path.join(output_dir, "review_decisions_template.csv")

        try:
            with (
                patch("generate_review_decisions_template.QUEUE_PATH", queue_path),
                patch("generate_review_decisions_template.OUTPUT_PATH", csv_path),
                patch(
                    "generate_review_decisions_template.OUTPUT_XLSX_PATH", xlsx_path
                ),
                patch("sys.stdout", StringIO()),
            ):
                result = main(["--dry-run"])

            self.assertFalse(
                os.path.exists(xlsx_path),
                "dry-run não deve criar o arquivo .xlsx",
            )
            self.assertFalse(
                os.path.exists(csv_path),
                "dry-run não deve criar o arquivo .csv",
            )
            self.assertEqual(result, 0)
        finally:
            os.unlink(queue_path)

    def test_main_normal_cria_xlsx(self):
        queue_json = _make_queue_json(copy.deepcopy(ALL_ITEMS))
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as fh:
            json.dump(queue_json, fh, ensure_ascii=False)
            queue_path = fh.name

        output_dir = tempfile.mkdtemp()
        xlsx_path = os.path.join(output_dir, "review_decisions_template.xlsx")
        csv_path = os.path.join(output_dir, "review_decisions_template.csv")

        try:
            with (
                patch("generate_review_decisions_template.QUEUE_PATH", queue_path),
                patch("generate_review_decisions_template.OUTPUT_PATH", csv_path),
                patch(
                    "generate_review_decisions_template.OUTPUT_XLSX_PATH", xlsx_path
                ),
                patch("sys.stdout", StringIO()),
            ):
                result = main([])

            self.assertTrue(os.path.exists(xlsx_path), ".xlsx deve ser criado em modo normal")
            self.assertTrue(os.path.exists(csv_path), ".csv deve ser criado em modo normal")
            self.assertEqual(result, 0)
        finally:
            os.unlink(queue_path)
            if os.path.exists(xlsx_path):
                os.unlink(xlsx_path)
            if os.path.exists(csv_path):
                os.unlink(csv_path)


if __name__ == "__main__":
    unittest.main()
