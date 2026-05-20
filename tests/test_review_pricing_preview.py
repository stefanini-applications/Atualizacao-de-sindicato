"""Testes para os comandos review-pricing-preview e generate-pricing-application-base.

Cobre os critérios de aceitação da US-PRJ-13:
  AC1   — review-pricing-preview exibe relatório completo (total, status, decisao, instrução)
  AC1-a — coluna decisao_aplicacao adicionada vazia quando ausente
  AC2   — generate-pricing-application-base filtra duplo critério; header-only + saída 1 quando vazio
  AC3   — arquivo de prévia ausente → erro claro + saída não-zero + sem arquivo de saída
  AC4   — colunas de saída = originais de pricing + 9 de enriquecimento (incluindo decisao_aplicacao)
  AC5   — base_pricing.xlsx não é escrita por nenhum dos dois comandos
  AC6   — linhas aprovadas com status inelegível geram aviso explícito
"""

import os
import uuid
from io import StringIO
from pathlib import Path
from typing import List
from unittest.mock import patch

import openpyxl
import pytest

from src.services.pricing_preview_reviewer import (
    filtrar_aprovadas_elegiveis,
    garantir_coluna_decisao_aplicacao,
    ler_preview_xlsx,
    salvar_base_aplicacoes,
    COLUNA_DECISAO,
    STATUS_ELEGIVEL,
    VALOR_APROVADO,
)
from src.reports.preview_pricing import (
    imprimir_relatorio_revisao_preview,
    _STATUS_ORDEM,
)
from src.services.preview_writer import _COLUNAS_ADICIONADAS


# ── constantes e helpers ─────────────────────────────────────────────────────

_COLUNAS_PRICING = ["uf", "sindicato", "ano_referencia", "salario_base", "cargo"]
_COLUNAS_PREVIEW = _COLUNAS_PRICING + _COLUNAS_ADICIONADAS  # 14 colunas totais


def _criar_preview_xlsx(
    path: Path,
    linhas: List[dict],
    colunas: List[str] = None,
    incluir_decisao: bool = True,
) -> None:
    """Cria um arquivo preview_atualizacao_pricing.xlsx de teste."""
    cols = colunas if colunas is not None else (
        _COLUNAS_PREVIEW if incluir_decisao else
        [c for c in _COLUNAS_PREVIEW if c != COLUNA_DECISAO]
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(cols)
    for linha in linhas:
        ws.append([linha.get(c) for c in cols])
    wb.save(path)


def _linha(
    status: str = STATUS_ELEGIVEL,
    decisao: str = None,
    uf: str = "SP",
    sindicato: str = "Sind",
    ano: str = "2025",
) -> dict:
    """Cria uma linha de prévia de pricing para testes."""
    row = {
        "uf": uf,
        "sindicato": sindicato,
        "ano_referencia": ano,
        "salario_base": 3000,
        "cargo": "Analista",
        "id_registro_reajuste": str(uuid.uuid4()),
        "percentual_reajuste_final": "5%",
        "data_base_final": "2025-05-01",
        "vigencia_inicio_final": "2025-05-01",
        "vigencia_fim_final": "2026-04-30",
        "fonte_documento": "a.pdf",
        "status_aplicacao": status,
        "decisao_aplicacao": decisao,
        "observacao_aplicacao": None,
    }
    return row


# ── AC1-a: garantir_coluna_decisao_aplicacao ──────────────────────────────────

def test_ac1a_adiciona_coluna_quando_ausente(tmp_path):
    path = tmp_path / "preview.xlsx"
    colunas_sem_decisao = [c for c in _COLUNAS_PREVIEW if c != COLUNA_DECISAO]
    _criar_preview_xlsx(path, [_linha()], colunas=colunas_sem_decisao)

    adicionada = garantir_coluna_decisao_aplicacao(path)

    assert adicionada is True
    _, colunas = ler_preview_xlsx(path)
    assert COLUNA_DECISAO in colunas


def test_ac1a_coluna_existente_nao_modifica_arquivo(tmp_path):
    path = tmp_path / "preview.xlsx"
    _criar_preview_xlsx(path, [_linha(decisao="aprovado")])
    mtime_antes = path.stat().st_mtime

    adicionada = garantir_coluna_decisao_aplicacao(path)

    assert adicionada is False
    assert path.stat().st_mtime == mtime_antes


def test_ac1a_coluna_adicionada_com_celulas_em_branco(tmp_path):
    path = tmp_path / "preview.xlsx"
    colunas_sem_decisao = [c for c in _COLUNAS_PREVIEW if c != COLUNA_DECISAO]
    linhas = [_linha() for _ in range(3)]
    _criar_preview_xlsx(path, linhas, colunas=colunas_sem_decisao)

    garantir_coluna_decisao_aplicacao(path)

    rows, colunas = ler_preview_xlsx(path)
    idx = colunas.index(COLUNA_DECISAO)
    for row in rows:
        assert row.get(COLUNA_DECISAO) is None or row.get(COLUNA_DECISAO) == ""


# ── AC1: ler_preview_xlsx e relatório ────────────────────────────────────────

def test_ac1_ler_preview_retorna_linhas_e_colunas(tmp_path):
    path = tmp_path / "preview.xlsx"
    linhas_in = [_linha(status="reajuste_encontrado"), _linha(status="sem_correspondencia")]
    _criar_preview_xlsx(path, linhas_in)

    linhas, colunas = ler_preview_xlsx(path)

    assert len(linhas) == 2
    assert colunas == _COLUNAS_PREVIEW


def test_ac1_relatorio_exibe_distribuicao_status(tmp_path, capsys):
    contagens_status = {"reajuste_encontrado": 3, "sem_correspondencia": 1}
    contagens_decisao = {"aprovado": 2, "rejeitado": 0, "": 2}

    imprimir_relatorio_revisao_preview(4, contagens_status, contagens_decisao, False)

    out = capsys.readouterr().out
    assert "reajuste_encontrado" in out
    assert "sem_correspondencia" in out
    assert "3" in out
    assert "1" in out


def test_ac1_relatorio_exibe_distribuicao_decisao(tmp_path, capsys):
    contagens_status = {"reajuste_encontrado": 2}
    contagens_decisao = {"aprovado": 1, "rejeitado": 1, "": 0}

    imprimir_relatorio_revisao_preview(2, contagens_status, contagens_decisao, False)

    out = capsys.readouterr().out
    assert "aprovado" in out
    assert "rejeitado" in out


def test_ac1_relatorio_exibe_instrucao_ao_operador(capsys):
    imprimir_relatorio_revisao_preview(0, {}, {"": 0}, False)

    out = capsys.readouterr().out
    assert "decisao_aplicacao" in out
    assert "generate-pricing-application-base" in out


def test_ac1_relatorio_informa_coluna_adicionada(capsys):
    imprimir_relatorio_revisao_preview(1, {}, {"": 1}, coluna_adicionada=True)

    out = capsys.readouterr().out
    assert "adicionada" in out.lower()


def test_ac1_todos_cinco_status_exibidos(capsys):
    imprimir_relatorio_revisao_preview(0, {}, {}, False)

    out = capsys.readouterr().out
    for status in _STATUS_ORDEM:
        assert status in out


# ── AC2: filtrar_aprovadas_elegiveis ─────────────────────────────────────────

def test_ac2_inclui_linha_aprovada_elegivel():
    linhas = [_linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO)]
    aprovadas, inelegiveis = filtrar_aprovadas_elegiveis(linhas)
    assert len(aprovadas) == 1
    assert inelegiveis == 0


def test_ac2_exclui_linha_sem_decisao():
    linhas = [_linha(status=STATUS_ELEGIVEL, decisao=None)]
    aprovadas, inelegiveis = filtrar_aprovadas_elegiveis(linhas)
    assert len(aprovadas) == 0
    assert inelegiveis == 0


def test_ac2_exclui_linha_rejeitada():
    linhas = [_linha(status=STATUS_ELEGIVEL, decisao="rejeitado")]
    aprovadas, inelegiveis = filtrar_aprovadas_elegiveis(linhas)
    assert len(aprovadas) == 0
    assert inelegiveis == 0


def test_ac2_filtro_duplo_exige_ambos_criterios():
    """Linha com decisao=aprovado mas status inelegível não vai para aprovadas."""
    linhas = [_linha(status="sem_correspondencia", decisao=VALOR_APROVADO)]
    aprovadas, inelegiveis = filtrar_aprovadas_elegiveis(linhas)
    assert len(aprovadas) == 0
    assert inelegiveis == 1


def test_ac2_mistura_de_linhas():
    linhas = [
        _linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO),   # inclui
        _linha(status=STATUS_ELEGIVEL, decisao=None),              # exclui
        _linha(status="sem_correspondencia", decisao=VALOR_APROVADO),  # inelegível
        _linha(status="dados_insuficientes", decisao="rejeitado"), # exclui
    ]
    aprovadas, inelegiveis = filtrar_aprovadas_elegiveis(linhas)
    assert len(aprovadas) == 1
    assert inelegiveis == 1


def test_ac2_arquivo_saida_header_only_quando_sem_linhas(tmp_path):
    output_path = tmp_path / "aplicacoes.xlsx"
    salvar_base_aplicacoes(output_path, [], _COLUNAS_PREVIEW)

    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1  # apenas cabeçalho
    assert list(rows[0]) == _COLUNAS_PREVIEW


# ── AC3: arquivo de prévia ausente ───────────────────────────────────────────

def test_ac3_review_preview_ausente_retorna_erro(tmp_path):
    from src.cli import cmd_review_pricing_preview
    import argparse

    args = argparse.Namespace(preview=str(tmp_path / "nao_existe.xlsx"))
    rc = cmd_review_pricing_preview(args)
    assert rc != 0


def test_ac3_review_preview_ausente_nao_cria_arquivo(tmp_path):
    from src.cli import cmd_review_pricing_preview
    import argparse

    saida = tmp_path / "nao_existe.xlsx"
    args = argparse.Namespace(preview=str(saida))
    cmd_review_pricing_preview(args)
    assert not saida.exists()


def test_ac3_generate_base_preview_ausente_retorna_erro(tmp_path):
    from src.cli import cmd_generate_pricing_application_base
    import argparse

    output = tmp_path / "out.xlsx"
    args = argparse.Namespace(
        preview=str(tmp_path / "nao_existe.xlsx"),
        output=str(output),
    )
    rc = cmd_generate_pricing_application_base(args)
    assert rc != 0
    assert not output.exists()


def test_ac3_mensagem_orienta_executar_preview_pricing_update(tmp_path, capsys):
    from src.cli import cmd_review_pricing_preview
    import argparse

    args = argparse.Namespace(preview=str(tmp_path / "nao_existe.xlsx"))
    cmd_review_pricing_preview(args)

    err = capsys.readouterr().err
    assert "preview-pricing-update" in err


# ── AC4: colunas de saída ────────────────────────────────────────────────────

def test_ac4_nove_colunas_enriquecimento():
    """_COLUNAS_ADICIONADAS deve conter exatamente 9 colunas incluindo decisao_aplicacao."""
    assert len(_COLUNAS_ADICIONADAS) == 9
    assert COLUNA_DECISAO in _COLUNAS_ADICIONADAS


def test_ac4_ordem_colunas_enriquecimento():
    esperadas = [
        "id_registro_reajuste",
        "percentual_reajuste_final",
        "data_base_final",
        "vigencia_inicio_final",
        "vigencia_fim_final",
        "fonte_documento",
        "status_aplicacao",
        "decisao_aplicacao",
        "observacao_aplicacao",
    ]
    assert _COLUNAS_ADICIONADAS == esperadas


def test_ac4_arquivo_saida_contem_todas_colunas(tmp_path):
    output_path = tmp_path / "aplicacoes.xlsx"
    linha = _linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO)
    salvar_base_aplicacoes(output_path, [linha], _COLUNAS_PREVIEW)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert cabecalho == _COLUNAS_PREVIEW


def test_ac4_decisao_aplicacao_presente_no_arquivo_de_saida(tmp_path):
    output_path = tmp_path / "aplicacoes.xlsx"
    linha = _linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO)
    salvar_base_aplicacoes(output_path, [linha], _COLUNAS_PREVIEW)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert COLUNA_DECISAO in cabecalho


# ── AC5: base_pricing.xlsx não é tocada ──────────────────────────────────────

def test_ac5_review_nao_abre_base_pricing(tmp_path):
    from src.cli import cmd_review_pricing_preview
    import argparse

    preview_path = tmp_path / "preview.xlsx"
    _criar_preview_xlsx(preview_path, [_linha()])

    pricing_path = tmp_path / "base_pricing.xlsx"
    pricing_path.write_bytes(b"dummy")
    mtime_antes = pricing_path.stat().st_mtime

    args = argparse.Namespace(preview=str(preview_path))
    cmd_review_pricing_preview(args)

    assert pricing_path.stat().st_mtime == mtime_antes


def test_ac5_generate_nao_abre_base_pricing(tmp_path):
    from src.cli import cmd_generate_pricing_application_base
    import argparse

    preview_path = tmp_path / "preview.xlsx"
    _criar_preview_xlsx(preview_path, [_linha(decisao=VALOR_APROVADO)])

    pricing_path = tmp_path / "base_pricing.xlsx"
    pricing_path.write_bytes(b"dummy")
    mtime_antes = pricing_path.stat().st_mtime

    output_path = tmp_path / "out.xlsx"
    args = argparse.Namespace(preview=str(preview_path), output=str(output_path))
    cmd_generate_pricing_application_base(args)

    assert pricing_path.stat().st_mtime == mtime_antes


# ── AC6: aprovadas inelegíveis emitem aviso ───────────────────────────────────

def test_ac6_aviso_para_aprovadas_inelegiveis():
    inelegiveis_status = [
        "sem_correspondencia",
        "multiplas_correspondencias",
        "dados_insuficientes",
        "erro_aplicacao",
    ]
    linhas = [_linha(status=s, decisao=VALOR_APROVADO) for s in inelegiveis_status]
    aprovadas, total_inelegiveis = filtrar_aprovadas_elegiveis(linhas)

    assert len(aprovadas) == 0
    assert total_inelegiveis == len(inelegiveis_status)


def test_ac6_generate_emite_aviso_stderr_para_inelegiveis(tmp_path, capsys):
    from src.cli import cmd_generate_pricing_application_base
    import argparse

    preview_path = tmp_path / "preview.xlsx"
    linhas = [
        _linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO),         # elegível
        _linha(status="sem_correspondencia", decisao=VALOR_APROVADO),   # inelegível
    ]
    _criar_preview_xlsx(preview_path, linhas)

    output_path = tmp_path / "out.xlsx"
    args = argparse.Namespace(preview=str(preview_path), output=str(output_path))
    rc = cmd_generate_pricing_application_base(args)

    err = capsys.readouterr().err
    assert "inelegível" in err or "inelegivel" in err.lower() or "ignorad" in err.lower()
    assert rc == 0  # há linhas válidas, então retorna 0


def test_ac6_generate_saida_zero_quando_ha_linhas_validas(tmp_path):
    from src.cli import cmd_generate_pricing_application_base
    import argparse

    preview_path = tmp_path / "preview.xlsx"
    linhas = [_linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO)]
    _criar_preview_xlsx(preview_path, linhas)

    output_path = tmp_path / "out.xlsx"
    args = argparse.Namespace(preview=str(preview_path), output=str(output_path))
    rc = cmd_generate_pricing_application_base(args)

    assert rc == 0
    assert output_path.exists()


def test_ac6_generate_saida_nao_zero_quando_sem_linhas_validas(tmp_path):
    from src.cli import cmd_generate_pricing_application_base
    import argparse

    preview_path = tmp_path / "preview.xlsx"
    linhas = [_linha(status="sem_correspondencia", decisao=VALOR_APROVADO)]
    _criar_preview_xlsx(preview_path, linhas)

    output_path = tmp_path / "out.xlsx"
    args = argparse.Namespace(preview=str(preview_path), output=str(output_path))
    rc = cmd_generate_pricing_application_base(args)

    assert rc != 0


# ── integração: CLI end-to-end ────────────────────────────────────────────────

def test_cli_review_pricing_preview_exit0(tmp_path):
    """review-pricing-preview deve encerrar com código 0 quando prévia existe."""
    from src.cli import main

    preview_path = tmp_path / "data" / "preview_atualizacao_pricing.xlsx"
    preview_path.parent.mkdir(parents=True)
    _criar_preview_xlsx(preview_path, [_linha(decisao=VALOR_APROVADO)])

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        rc = main(["review-pricing-preview"])

    assert rc == 0


def test_cli_generate_pricing_application_base_exit0(tmp_path):
    """generate-pricing-application-base deve encerrar com código 0 quando há linhas aptas."""
    from src.cli import main

    preview_path = tmp_path / "data" / "preview_atualizacao_pricing.xlsx"
    preview_path.parent.mkdir(parents=True)
    linhas = [_linha(status=STATUS_ELEGIVEL, decisao=VALOR_APROVADO)]
    _criar_preview_xlsx(preview_path, linhas)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        rc = main(["generate-pricing-application-base"])

    assert rc == 0
    output = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    assert output.exists()

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 2  # cabeçalho + 1 linha


def test_cli_generate_cria_header_only_quando_sem_linhas_aptas(tmp_path):
    """generate-pricing-application-base cria xlsx com apenas cabeçalho quando não há linhas aptas."""
    from src.cli import main

    preview_path = tmp_path / "data" / "preview_atualizacao_pricing.xlsx"
    preview_path.parent.mkdir(parents=True)
    linhas = [_linha(status=STATUS_ELEGIVEL, decisao=None)]  # sem decisao
    _criar_preview_xlsx(preview_path, linhas)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        rc = main(["generate-pricing-application-base"])

    assert rc != 0
    output = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    assert output.exists()

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1  # apenas cabeçalho
