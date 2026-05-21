"""Testes para o comando apply-pricing-updates (US-PRJ-14).

Cobre os critérios de aceitação:
  AC1  — arquivo de saída contém todas as colunas originais + 7 novos campos,
         mesmo número de linhas que a base original
  AC2  — correspondência única válida → atualizado com campos corretos
  AC3  — sem correspondência → nao_atualizado com campos corretos
  AC4  — base_pricing.xlsx permanece byte a byte idêntica após execução
  AC5  — id_registro_reajuste e data_hora_aplicacao preenchidos em linhas atualizado
  AC6  — --value-column obrigatório; ausência encerra com erro sem gerar arquivo
  AC7  — coluna ausente/vazia/não-numérica → erro_atualizacao com motivo específico
  AC8  — múltiplos registros aprovados → erro_atualizacao com mensagem padrão
  AC9  — apenas valores de status permitidos no arquivo de saída
  AC10 — parsing de percentual em múltiplos formatos
"""

import argparse
import uuid
from io import StringIO
from pathlib import Path
from typing import List
from unittest.mock import patch

import openpyxl
import pytest

from src.services.pricing_applier import (
    aplicar_reajustes,
    _parsear_percentual,
    STATUS_ATUALIZADO,
    STATUS_NAO_ATUALIZADO,
    STATUS_ERRO,
    OBS_NAO_ATUALIZADO,
    OBS_MULTIPLOS,
    COLUNAS_SAIDA,
)
from src.services.updated_pricing_writer import salvar_base_atualizada


# ── helpers ───────────────────────────────────────────────────────────────────

_COLUNAS_PRICING = ["uf", "sindicato", "ano_referencia", "valor_pricing", "cargo"]

_COLUNAS_APROVACOES = _COLUNAS_PRICING + [
    "id_registro_reajuste",
    "percentual_reajuste_final",
    "data_base_final",
    "vigencia_inicio_final",
    "vigencia_fim_final",
    "fonte_documento",
    "status_aplicacao",
    "decisao_aplicacao",
    "observacao_aplicacao",
]

_TIMESTAMP = "2025-05-20T10:00:00+00:00"


def _linha_pricing(
    uf: str = "SP",
    sindicato: str = "Sind A",
    ano: str = "2025",
    valor: object = 3000.0,
    cargo: str = "Analista",
) -> dict:
    return {
        "uf": uf,
        "sindicato": sindicato,
        "ano_referencia": ano,
        "valor_pricing": valor,
        "cargo": cargo,
    }


def _linha_aprovada(
    uf: str = "SP",
    sindicato: str = "Sind A",
    ano: str = "2025",
    id_registro: str = None,
    percentual: object = "5%",
) -> dict:
    return {
        "uf": uf,
        "sindicato": sindicato,
        "ano_referencia": ano,
        "valor_pricing": 3000.0,
        "cargo": "Analista",
        "id_registro_reajuste": id_registro or str(uuid.uuid4()),
        "percentual_reajuste_final": percentual,
        "data_base_final": "2025-05-01",
        "vigencia_inicio_final": "2025-05-01",
        "vigencia_fim_final": "2026-04-30",
        "fonte_documento": "doc.pdf",
        "status_aplicacao": "reajuste_encontrado",
        "decisao_aplicacao": "aprovado",
        "observacao_aplicacao": None,
    }


def _criar_xlsx(path: Path, linhas: List[dict], colunas: List[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(colunas)
    for linha in linhas:
        ws.append([linha.get(c) for c in colunas])
    wb.save(path)


# ── AC1: colunas de saída e contagem de linhas ───────────────────────────────

def test_ac1_arquivo_saida_contem_colunas_originais_mais_7(tmp_path):
    linhas = [_linha_pricing()]
    salvar_base_atualizada(
        tmp_path / "saida.xlsx",
        [aplicar_reajustes(linhas, [], "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP)[0]],
        _COLUNAS_PRICING,
    )
    wb = openpyxl.load_workbook(tmp_path / "saida.xlsx")
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert cabecalho == _COLUNAS_PRICING + COLUNAS_SAIDA


def test_ac1_numero_linhas_identico_base_original(tmp_path):
    n = 5
    linhas = [_linha_pricing(sindicato=f"S{i}") for i in range(n)]
    enriquecidas = aplicar_reajustes(
        linhas, [], "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert len(enriquecidas) == n

    output = tmp_path / "saida.xlsx"
    salvar_base_atualizada(output, enriquecidas, _COLUNAS_PRICING)
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == n + 1  # cabeçalho + n linhas


def test_ac1_sete_colunas_novas_exatas():
    esperadas = [
        "valor_original",
        "percentual_reajuste_aplicado",
        "valor_reajustado",
        "id_registro_reajuste",
        "data_hora_aplicacao",
        "status_atualizacao",
        "observacao_atualizacao",
    ]
    assert COLUNAS_SAIDA == esperadas


# ── AC2: correspondência única válida → atualizado ───────────────────────────

def test_ac2_valor_reajustado_calculado_corretamente():
    linhas = [_linha_pricing(valor=1000.0)]
    aprovacoes = [_linha_aprovada(percentual="10%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_ATUALIZADO
    assert linha["valor_original"] == 1000.0
    assert linha["valor_reajustado"] == pytest.approx(1100.0)
    assert linha["percentual_reajuste_aplicado"] == pytest.approx(10.0)


def test_ac2_valor_original_preservado():
    linhas = [_linha_pricing(valor=2500.0)]
    aprovacoes = [_linha_aprovada(percentual="5")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["valor_original"] == 2500.0


def test_ac2_observacao_vazia_quando_atualizado():
    linhas = [_linha_pricing()]
    aprovacoes = [_linha_aprovada(percentual="5%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["observacao_atualizacao"] is None


def test_ac2_dados_originais_preservados_nos_outros_campos():
    linhas = [_linha_pricing(cargo="Engenheiro", valor=5000.0)]
    aprovacoes = [_linha_aprovada(percentual="3%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["cargo"] == "Engenheiro"
    assert resultado[0]["uf"] == "SP"


# ── AC3: sem correspondência → nao_atualizado ────────────────────────────────

def test_ac3_status_nao_atualizado_quando_sem_correspondencia():
    linhas = [_linha_pricing(uf="RJ")]
    aprovacoes = [_linha_aprovada(uf="SP")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_NAO_ATUALIZADO


def test_ac3_valor_reajustado_igual_original_quando_nao_atualizado():
    linhas = [_linha_pricing(valor=4000.0, uf="MG")]
    aprovacoes = [_linha_aprovada(uf="SP")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["valor_original"] == 4000.0
    assert linha["valor_reajustado"] == 4000.0


def test_ac3_campos_em_branco_quando_nao_atualizado():
    linhas = [_linha_pricing(uf="MG")]
    aprovacoes = [_linha_aprovada(uf="SP")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["percentual_reajuste_aplicado"] is None
    assert linha["id_registro_reajuste"] is None
    assert linha["data_hora_aplicacao"] is None


def test_ac3_observacao_mensagem_padrao_quando_nao_atualizado():
    linhas = [_linha_pricing(uf="MG")]
    aprovacoes = []
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["observacao_atualizacao"] == OBS_NAO_ATUALIZADO


# ── AC4: base_pricing.xlsx inalterada ─────────────────────────────────────────

def test_ac4_base_pricing_nao_alterada(tmp_path):
    from src.cli import cmd_apply_pricing_updates

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing()], _COLUNAS_PRICING)
    conteudo_antes = pricing_path.read_bytes()

    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada()], _COLUNAS_APROVACOES)

    output_path = tmp_path / "data" / "base_pricing_atualizada.xlsx"

    args = argparse.Namespace(
        value_column="valor_pricing",
        pricing=str(pricing_path),
        approvals=str(aprovacoes_path),
        output=str(output_path),
    )
    with patch("src.cli._raiz_repo", return_value=Path("/")):
        args.pricing = str(pricing_path)
        args.approvals = str(aprovacoes_path)
        args.output = str(output_path)
        cmd_apply_pricing_updates(args)

    assert pricing_path.read_bytes() == conteudo_antes


# ── AC5: id_registro_reajuste e data_hora_aplicacao preenchidos ───────────────

def test_ac5_id_registro_preenchido_em_linha_atualizada():
    id_esperado = str(uuid.uuid4())
    linhas = [_linha_pricing()]
    aprovacoes = [_linha_aprovada(id_registro=id_esperado, percentual="5%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["id_registro_reajuste"] == id_esperado


def test_ac5_data_hora_aplicacao_preenchida_em_linha_atualizada():
    linhas = [_linha_pricing()]
    aprovacoes = [_linha_aprovada(percentual="5%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["data_hora_aplicacao"] == _TIMESTAMP


def test_ac5_id_e_timestamp_vazios_em_linha_nao_atualizada():
    linhas = [_linha_pricing(uf="MG")]
    aprovacoes = [_linha_aprovada(uf="SP")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["id_registro_reajuste"] is None
    assert resultado[0]["data_hora_aplicacao"] is None


# ── AC6: --value-column obrigatório ──────────────────────────────────────────

def test_ac6_ausencia_value_column_encerra_com_erro(tmp_path, capsys):
    from src.cli import main

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing()], _COLUNAS_PRICING)

    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada()], _COLUNAS_APROVACOES)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        with pytest.raises(SystemExit) as exc_info:
            main(["apply-pricing-updates"])

    assert exc_info.value.code != 0


def test_ac6_sem_value_column_nao_gera_arquivo(tmp_path):
    from src.cli import main

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing()], _COLUNAS_PRICING)

    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada()], _COLUNAS_APROVACOES)

    output_path = tmp_path / "data" / "base_pricing_atualizada.xlsx"

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        with pytest.raises(SystemExit):
            main(["apply-pricing-updates"])

    assert not output_path.exists()


# ── AC7: validação da coluna por linha ────────────────────────────────────────

def test_ac7_coluna_nao_encontrada():
    linhas = [{"uf": "SP", "sindicato": "S", "ano_referencia": "2025", "outro": 100}]
    aprovacoes = [_linha_aprovada()]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_ERRO
    assert "não encontrada" in linha["observacao_atualizacao"]
    assert "valor_pricing" in linha["observacao_atualizacao"]


def test_ac7_valor_ausente_na_coluna():
    linhas = [_linha_pricing(valor=None)]
    aprovacoes = [_linha_aprovada()]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_ERRO
    assert "ausente" in linha["observacao_atualizacao"]


def test_ac7_valor_nao_numerico():
    linhas = [_linha_pricing(valor="abc")]
    aprovacoes = [_linha_aprovada()]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_ERRO
    assert "não numérico" in linha["observacao_atualizacao"]
    assert "abc" in linha["observacao_atualizacao"]


def test_ac7_erro_coluna_nao_interrompe_outras_linhas():
    linhas = [
        _linha_pricing(valor="invalido", sindicato="S1"),
        _linha_pricing(valor=1000.0, sindicato="S2"),
    ]
    aprovacoes = [_linha_aprovada(sindicato="S2", percentual="5%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert len(resultado) == 2
    assert resultado[0]["status_atualizacao"] == STATUS_ERRO
    assert resultado[1]["status_atualizacao"] == STATUS_ATUALIZADO


def test_ac7_valor_string_vazia():
    linhas = [_linha_pricing(valor="")]
    aprovacoes = [_linha_aprovada()]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["status_atualizacao"] == STATUS_ERRO
    assert "ausente" in resultado[0]["observacao_atualizacao"]


# ── AC8: múltiplos registros aprovados ────────────────────────────────────────

def test_ac8_multiplos_registros_marca_erro():
    linhas = [_linha_pricing()]
    aprovacoes = [
        _linha_aprovada(percentual="5%"),
        _linha_aprovada(percentual="10%"),
    ]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_ERRO
    assert linha["observacao_atualizacao"] == OBS_MULTIPLOS


def test_ac8_multiplos_registros_nao_aplica_reajuste():
    linhas = [_linha_pricing(valor=1000.0)]
    aprovacoes = [
        _linha_aprovada(percentual="5%"),
        _linha_aprovada(percentual="10%"),
    ]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["valor_reajustado"] is None
    assert linha["percentual_reajuste_aplicado"] is None


# ── AC9: status_atualizacao apenas valores permitidos ────────────────────────

def test_ac9_apenas_status_permitidos():
    valores_permitidos = {STATUS_ATUALIZADO, STATUS_NAO_ATUALIZADO, STATUS_ERRO}
    linhas = [
        _linha_pricing(sindicato="S1"),          # vai ser atualizado
        _linha_pricing(sindicato="S2", uf="RJ"), # sem correspondência
        _linha_pricing(sindicato="S3", valor="x"), # erro coluna
    ]
    aprovacoes = [_linha_aprovada(sindicato="S1", percentual="5%")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    for linha in resultado:
        assert linha["status_atualizacao"] in valores_permitidos


# ── AC10: parsing de percentual ───────────────────────────────────────────────

@pytest.mark.parametrize("valor,esperado", [
    ("5",    5.0),
    ("5.0",  5.0),
    ("5,0",  5.0),
    ("5%",   5.0),
    ("5,0%", 5.0),
    ("5.0%", 5.0),
    ("-3%",  -3.0),
    ("10.5", 10.5),
])
def test_ac10_parsing_percentual_formatos_suportados(valor, esperado):
    resultado, erro = _parsear_percentual(valor)
    assert erro is None
    assert resultado == pytest.approx(esperado)


def test_ac10_parsing_none_retorna_erro():
    resultado, erro = _parsear_percentual(None)
    assert resultado is None
    assert erro is not None
    assert "None" in erro


def test_ac10_parsing_string_invalida_retorna_erro():
    resultado, erro = _parsear_percentual("abc%")
    assert resultado is None
    assert erro is not None
    assert "abc%" in erro


def test_ac10_erro_percentual_marca_erro_na_linha():
    linhas = [_linha_pricing()]
    aprovacoes = [_linha_aprovada(percentual="invalido")]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    linha = resultado[0]
    assert linha["status_atualizacao"] == STATUS_ERRO
    assert "percentual_reajuste_final" in linha["observacao_atualizacao"]


def test_ac10_erro_percentual_nao_interrompe_outras_linhas():
    linhas = [
        _linha_pricing(sindicato="S1"),
        _linha_pricing(sindicato="S2"),
    ]
    aprovacoes = [
        _linha_aprovada(sindicato="S1", percentual="invalido"),
        _linha_aprovada(sindicato="S2", percentual="5%"),
    ]
    resultado = aplicar_reajustes(
        linhas, aprovacoes, "uf", "sindicato", "ano_referencia", "valor_pricing", _TIMESTAMP
    )
    assert resultado[0]["status_atualizacao"] == STATUS_ERRO
    assert resultado[1]["status_atualizacao"] == STATUS_ATUALIZADO


# ── integração CLI end-to-end ─────────────────────────────────────────────────

def test_cli_apply_pricing_updates_exit0(tmp_path):
    from src.cli import main

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing()], _COLUNAS_PRICING)

    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada()], _COLUNAS_APROVACOES)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        rc = main(["apply-pricing-updates", "--value-column", "valor_pricing"])

    assert rc == 0
    output = tmp_path / "data" / "base_pricing_atualizada.xlsx"
    assert output.exists()


def test_cli_apply_pricing_updates_colunas_corretas(tmp_path):
    from src.cli import main

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing()], _COLUNAS_PRICING)

    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada()], _COLUNAS_APROVACOES)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        main(["apply-pricing-updates", "--value-column", "valor_pricing"])

    output = tmp_path / "data" / "base_pricing_atualizada.xlsx"
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert cabecalho == _COLUNAS_PRICING + COLUNAS_SAIDA


def test_cli_apply_pricing_updates_arquivo_ausente_retorna_erro(tmp_path, capsys):
    from src.cli import main

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        rc = main(["apply-pricing-updates", "--value-column", "valor_pricing"])

    assert rc != 0


def test_cli_apply_pricing_conteudo_linha_atualizada(tmp_path):
    """Verifica que a linha atualizada no arquivo tem os valores corretos."""
    from src.cli import main

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing(valor=2000.0)], _COLUNAS_PRICING)

    id_reg = str(uuid.uuid4())
    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada(id_registro=id_reg, percentual="10%")], _COLUNAS_APROVACOES)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        main(["apply-pricing-updates", "--value-column", "valor_pricing"])

    output = tmp_path / "data" / "base_pricing_atualizada.xlsx"
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    row = {cabecalho[i]: ws.cell(2, i + 1).value for i in range(len(cabecalho))}

    assert row["status_atualizacao"] == STATUS_ATUALIZADO
    assert row["valor_original"] == pytest.approx(2000.0)
    assert row["valor_reajustado"] == pytest.approx(2200.0)
    assert row["percentual_reajuste_aplicado"] == pytest.approx(10.0)
    assert row["id_registro_reajuste"] == id_reg


def test_cli_reajuste_negativo_suportado(tmp_path):
    """Reajustes negativos (reduções) são matematicamente suportados."""
    from src.cli import main

    pricing_path = tmp_path / "data" / "base_pricing.xlsx"
    pricing_path.parent.mkdir(parents=True)
    _criar_xlsx(pricing_path, [_linha_pricing(valor=1000.0)], _COLUNAS_PRICING)

    aprovacoes_path = tmp_path / "data" / "aplicacoes_pricing_aprovadas.xlsx"
    _criar_xlsx(aprovacoes_path, [_linha_aprovada(percentual="-5%")], _COLUNAS_APROVACOES)

    with patch("src.cli._raiz_repo", return_value=tmp_path):
        main(["apply-pricing-updates", "--value-column", "valor_pricing"])

    output = tmp_path / "data" / "base_pricing_atualizada.xlsx"
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    row = {cabecalho[i]: ws.cell(2, i + 1).value for i in range(len(cabecalho))}

    assert row["status_atualizacao"] == STATUS_ATUALIZADO
    assert row["valor_reajustado"] == pytest.approx(950.0)
