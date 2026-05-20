"""Testes para o comando preview-pricing-update e módulos de prévia de pricing.

Cobre os critérios de aceitação da US-PRJ-12:
  AC1 — validação dos arquivos de entrada: ausente → mensagem + saída não-zero
  AC2 — preservação da base original (escrita atômica, original inalterado)
  AC3 — cruzamento e status por linha (regras de correspondência)
  AC4 — campos adicionados na prévia (8 colunas + originais preservadas)
  AC5 — relatório: totais por status com confirmação de soma
"""

import json
import os
import uuid
from pathlib import Path
from typing import List
from unittest.mock import patch

import openpyxl
import pytest

from src.models.reajuste_aprovado import ReajusteAprovado
from src.models.linha_preview_pricing import COLUNAS_PREVIEW, LinhaPreviewPricing
from src.services.pricing_reader import ler_base_pricing, ErroCabecalhoPricing
from src.services.pricing_preview import (
    gerar_preview,
    STATUS_ENCONTRADO,
    STATUS_SEM_CORRESPONDENCIA,
    STATUS_MULTIPLAS,
    STATUS_INSUFICIENTE,
    STATUS_ERRO,
)
from src.services.preview_writer import salvar_preview
from src.reports.preview_pricing import imprimir_relatorio_preview


# ── helpers ───────────────────────────────────────────────────────────────────

def _aprovado(
    uf: str = "SP",
    sindicato: str = "SindTest",
    ano_referencia: str = "2024",
    percentual: str = "5%",
    id_registro: str = None,
) -> ReajusteAprovado:
    return ReajusteAprovado(
        id_registro=id_registro or str(uuid.uuid4()),
        caminho="CCT/SP/SindTest/doc.pdf",
        nome_arquivo="doc.pdf",
        uf=uf,
        sindicato=sindicato,
        tipo_documento="CCT",
        ano_referencia=ano_referencia,
        tipo_clausula="reajuste_salarial",
        trecho_original="Reajuste de 5%.",
        percentual_reajuste_original=percentual,
        percentual_reajuste_final=percentual,
        data_base_original="2024-01-01",
        data_base_final="2024-01-01",
        vigencia_inicio_original="2024-01-01",
        vigencia_inicio_final="2024-01-01",
        vigencia_fim_original="2024-12-31",
        vigencia_fim_final="2024-12-31",
        status_validacao="aprovado",
        responsavel_validacao="operador",
        data_hora_validacao="2025-01-01T00:00:00+00:00",
        observacao_validacao=None,
        data_hora_geracao="2025-01-01T00:00:00+00:00",
    )


def _criar_xlsx(path: Path, headers: list, rows: list) -> None:
    """Cria um arquivo XLSX simples com cabeçalho + linhas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _criar_json_aprovados(path: Path, aprovados: List[ReajusteAprovado]) -> None:
    from src.services.approved_store import salvar_aprovados
    salvar_aprovados(path, aprovados)


# ── AC1: validação dos arquivos de entrada ────────────────────────────────────

def test_ac1_adjustments_ausente_retorna_1(tmp_path):
    from src.cli import main
    pricing = tmp_path / "pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato", "ano_referencia"], [["SP", "Sind", "2024"]])
    codigo = main([
        "preview-pricing-update",
        "--adjustments", str(tmp_path / "nao_existe.json"),
        "--pricing", str(pricing),
        "--output", str(tmp_path / "out.xlsx"),
    ])
    assert codigo == 1


def test_ac1_pricing_ausente_retorna_1(tmp_path):
    from src.cli import main
    adj = tmp_path / "adj.json"
    _criar_json_aprovados(adj, [_aprovado()])
    codigo = main([
        "preview-pricing-update",
        "--adjustments", str(adj),
        "--pricing", str(tmp_path / "nao_existe.xlsx"),
        "--output", str(tmp_path / "out.xlsx"),
    ])
    assert codigo == 1


def test_ac1_ambos_ausentes_retorna_1(tmp_path):
    from src.cli import main
    codigo = main([
        "preview-pricing-update",
        "--adjustments", str(tmp_path / "a.json"),
        "--pricing", str(tmp_path / "b.xlsx"),
        "--output", str(tmp_path / "out.xlsx"),
    ])
    assert codigo == 1


def test_ac1_arquivo_ausente_nao_cria_saida(tmp_path):
    from src.cli import main
    output = tmp_path / "out.xlsx"
    main([
        "preview-pricing-update",
        "--adjustments", str(tmp_path / "nao_existe.json"),
        "--pricing", str(tmp_path / "nao_existe.xlsx"),
        "--output", str(output),
    ])
    assert not output.exists()


def test_ac1_adjustments_ausente_menciona_arquivo_no_stderr(tmp_path, capsys):
    from src.cli import main
    pricing = tmp_path / "pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato"], [["SP", "Sind"]])
    main([
        "preview-pricing-update",
        "--adjustments", str(tmp_path / "nao_existe.json"),
        "--pricing", str(pricing),
        "--output", str(tmp_path / "out.xlsx"),
    ])
    captured = capsys.readouterr()
    assert "nao_existe.json" in captured.err


def test_ac1_pricing_ausente_menciona_arquivo_no_stderr(tmp_path, capsys):
    from src.cli import main
    adj = tmp_path / "adj.json"
    _criar_json_aprovados(adj, [_aprovado()])
    main([
        "preview-pricing-update",
        "--adjustments", str(adj),
        "--pricing", str(tmp_path / "nao_existe.xlsx"),
        "--output", str(tmp_path / "out.xlsx"),
    ])
    captured = capsys.readouterr()
    assert "nao_existe.xlsx" in captured.err


# ── AC2: preservação da base original ────────────────────────────────────────

def test_ac2_base_pricing_nao_modificada(tmp_path):
    """Após execução bem-sucedida o conteúdo do pricing original é inalterado."""
    from src.cli import main
    pricing = tmp_path / "pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato", "ano_referencia"], [["SP", "SindTest", "2024"]])
    conteudo_original = pricing.read_bytes()
    mtime_original = pricing.stat().st_mtime

    adj = tmp_path / "adj.json"
    _criar_json_aprovados(adj, [_aprovado()])

    main([
        "preview-pricing-update",
        "--adjustments", str(adj),
        "--pricing", str(pricing),
        "--output", str(tmp_path / "out.xlsx"),
    ])

    assert pricing.read_bytes() == conteudo_original
    assert pricing.stat().st_mtime == mtime_original


def test_ac2_escrita_atomica_sem_residuo(tmp_path):
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "Sind", "ano_referencia": "2024"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    output = tmp_path / "out.xlsx"
    salvar_preview(output, headers, linhas)

    arquivos = list(tmp_path.iterdir())
    assert output in arquivos
    assert [f for f in arquivos if ".tmp" in f.name] == []


def test_ac2_escrita_atomica_nao_corrompe_existente(tmp_path):
    """Se os.replace falhar, o arquivo de saída existente deve permanecer intacto."""
    headers = ["uf", "sindicato"]
    rows = [{"uf": "SP", "sindicato": "Sind"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])

    output = tmp_path / "out.xlsx"
    conteudo_inicial = b"conteudo original"
    output.write_bytes(conteudo_inicial)

    with patch("os.replace", side_effect=OSError("falha simulada")):
        with pytest.raises(OSError):
            salvar_preview(output, headers, linhas)

    assert output.read_bytes() == conteudo_inicial


def test_ac2_arquivo_temporario_removido_em_falha(tmp_path):
    headers = ["uf", "sindicato"]
    rows = [{"uf": "SP", "sindicato": "Sind"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    output = tmp_path / "out.xlsx"

    with patch("os.replace", side_effect=OSError("falha simulada")):
        with pytest.raises(OSError):
            salvar_preview(output, headers, linhas)

    tmp_files = [f for f in tmp_path.iterdir() if ".tmp" in f.name]
    assert tmp_files == []


# ── AC3: cruzamento e status por linha ───────────────────────────────────────

def test_ac3_reajuste_encontrado_match_exato():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "ano_referencia": "2024"}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_ENCONTRADO


def test_ac3_sem_correspondencia_sem_match():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "RJ", "sindicato": "OutroSind", "ano_referencia": "2024"}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_SEM_CORRESPONDENCIA


def test_ac3_multiplas_correspondencias():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "ano_referencia": "2024"}]
    id1 = str(uuid.uuid4())
    id2 = str(uuid.uuid4())
    aprovados = [
        _aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024", id_registro=id1),
        _aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024", id_registro=id2),
    ]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_MULTIPLAS
    assert id1 in linhas[0].observacao_aplicacao
    assert id2 in linhas[0].observacao_aplicacao


def test_ac3_dados_insuficientes_sem_uf():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": None, "sindicato": "SindTest", "ano_referencia": "2024"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    assert linhas[0].status_aplicacao == STATUS_INSUFICIENTE


def test_ac3_dados_insuficientes_sem_sindicato():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "", "ano_referencia": "2024"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    assert linhas[0].status_aplicacao == STATUS_INSUFICIENTE


def test_ac3_dados_insuficientes_sem_coluna_uf():
    """Quando a coluna 'uf' não existe na planilha, toda linha é dados_insuficientes."""
    headers = ["sindicato", "ano_referencia"]
    rows = [{"sindicato": "SindTest", "ano_referencia": "2024"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    assert linhas[0].status_aplicacao == STATUS_INSUFICIENTE


def test_ac3_dados_insuficientes_sem_coluna_sindicato():
    headers = ["uf", "ano_referencia"]
    rows = [{"uf": "SP", "ano_referencia": "2024"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    assert linhas[0].status_aplicacao == STATUS_INSUFICIENTE


def test_ac3_match_por_uf_sindicato_quando_sem_ano_referencia():
    """Sem coluna ano_referencia, cruzamento é feito só por uf+sindicato."""
    headers = ["uf", "sindicato"]
    rows = [{"uf": "SP", "sindicato": "SindTest"}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_ENCONTRADO


def test_ac3_match_case_insensitive():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "sp", "sindicato": "sindtest", "ano_referencia": "2024"}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_ENCONTRADO


def test_ac3_match_com_espacos_extras():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": " SP ", "sindicato": " SindTest ", "ano_referencia": " 2024 "}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_ENCONTRADO


def test_ac3_ano_numerico_normalizado():
    """Ano como número inteiro ou float deve ser normalizado para string."""
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "ano_referencia": 2024}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_ENCONTRADO


def test_ac3_ano_float_normalizado():
    """Ano como float 2024.0 deve ser normalizado para '2024'."""
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "ano_referencia": 2024.0}]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert linhas[0].status_aplicacao == STATUS_ENCONTRADO


def test_ac3_cada_linha_recebe_exatamente_um_status():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [
        {"uf": "SP", "sindicato": "SindTest", "ano_referencia": "2024"},  # encontrado
        {"uf": "RJ", "sindicato": "Outro", "ano_referencia": "2024"},     # sem correspondencia
        {"uf": None, "sindicato": "Sind", "ano_referencia": "2024"},      # dados insuficientes
    ]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    assert len(linhas) == len(rows)
    assert all(l.status_aplicacao for l in linhas)


# ── AC4: campos adicionados na prévia ────────────────────────────────────────

def test_ac4_colunas_originais_preservadas_no_xlsx(tmp_path):
    """O XLSX de saída deve conter todas as colunas originais da base de pricing."""
    headers = ["uf", "sindicato", "ano_referencia", "cargo", "salario_base"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "ano_referencia": "2024",
             "cargo": "Analista", "salario_base": 3000}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    output = tmp_path / "out.xlsx"
    salvar_preview(output, headers, linhas)

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    cabecalhos_saida = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    for h in headers:
        assert h in cabecalhos_saida


def test_ac4_oito_colunas_preview_adicionadas(tmp_path):
    """O XLSX de saída deve conter os 8 campos de prévia definidos em COLUNAS_PREVIEW."""
    headers = ["uf", "sindicato"]
    rows = [{"uf": "SP", "sindicato": "SindTest"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    output = tmp_path / "out.xlsx"
    salvar_preview(output, headers, linhas)

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    cabecalhos_saida = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    for campo in COLUNAS_PREVIEW:
        assert campo in cabecalhos_saida


def test_ac4_campos_preview_presentes_mesmo_sem_match(tmp_path):
    """Linhas sem correspondência ainda devem ter os 8 campos (com valor None/vazio)."""
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "RJ", "sindicato": "SemMatch", "ano_referencia": "2024"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    assert linhas[0].id_registro_reajuste is None
    assert linhas[0].percentual_reajuste_final is None
    assert linhas[0].status_aplicacao == STATUS_SEM_CORRESPONDENCIA


def test_ac4_campos_preenchidos_quando_encontrado():
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "ano_referencia": "2024"}]
    r = _aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024", percentual="7%")
    linhas = gerar_preview(headers, rows, [r])
    l = linhas[0]
    assert l.id_registro_reajuste == r.id_registro
    assert l.percentual_reajuste_final == "7%"
    assert l.data_base_final == r.data_base_final
    assert l.vigencia_inicio_final == r.vigencia_inicio_final
    assert l.vigencia_fim_final == r.vigencia_fim_final
    assert l.fonte_documento == "doc.pdf"


def test_ac4_dados_originais_preservados_na_linha():
    headers = ["uf", "sindicato", "cargo"]
    rows = [{"uf": "SP", "sindicato": "SindTest", "cargo": "Gerente"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    assert linhas[0].dados_originais["cargo"] == "Gerente"
    assert linhas[0].dados_originais["uf"] == "SP"


# ── AC5: relatório de simulação no terminal ───────────────────────────────────

def test_ac5_relatorio_exibe_total_avaliados(capsys):
    headers = ["uf", "sindicato"]
    rows = [
        {"uf": "SP", "sindicato": "SindTest"},
        {"uf": "RJ", "sindicato": "Outro"},
    ]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    imprimir_relatorio_preview(linhas)
    captured = capsys.readouterr()
    assert "2" in captured.out


def test_ac5_relatorio_exibe_totais_por_status(capsys):
    headers = ["uf", "sindicato", "ano_referencia"]
    rows = [
        {"uf": "SP", "sindicato": "SindTest", "ano_referencia": "2024"},
        {"uf": "RJ", "sindicato": "Outro", "ano_referencia": "2024"},
        {"uf": None, "sindicato": "Sind", "ano_referencia": "2024"},
    ]
    aprovados = [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")]
    linhas = gerar_preview(headers, rows, aprovados)
    imprimir_relatorio_preview(linhas)
    captured = capsys.readouterr()
    assert STATUS_ENCONTRADO in captured.out
    assert STATUS_SEM_CORRESPONDENCIA in captured.out
    assert STATUS_INSUFICIENTE in captured.out


def test_ac5_relatorio_confirma_soma_igual_total(capsys):
    headers = ["uf", "sindicato"]
    rows = [{"uf": "SP", "sindicato": "SindTest"}]
    linhas = gerar_preview(headers, rows, [_aprovado()])
    imprimir_relatorio_preview(linhas)
    captured = capsys.readouterr()
    assert "OK" in captured.out


# ── pricing_reader: leitura e validação ──────────────────────────────────────

def test_reader_lê_cabeçalhos_e_linhas(tmp_path):
    xlsx = tmp_path / "pricing.xlsx"
    _criar_xlsx(xlsx, ["uf", "sindicato", "cargo"], [["SP", "Sind", "Analista"]])
    headers, rows = ler_base_pricing(xlsx)
    assert headers == ["uf", "sindicato", "cargo"]
    assert len(rows) == 1
    assert rows[0]["uf"] == "SP"


def test_reader_celula_vazia_vira_none(tmp_path):
    xlsx = tmp_path / "pricing.xlsx"
    _criar_xlsx(xlsx, ["uf", "sindicato"], [["SP", None]])
    _, rows = ler_base_pricing(xlsx)
    assert rows[0]["sindicato"] is None


def test_reader_planilha_vazia_retorna_listas_vazias(tmp_path):
    wb = openpyxl.Workbook()
    wb.save(tmp_path / "empty.xlsx")
    headers, rows = ler_base_pricing(tmp_path / "empty.xlsx")
    assert headers == []
    assert rows == []


def test_reader_cabecalho_em_branco_levanta_erro(tmp_path):
    xlsx = tmp_path / "pricing.xlsx"
    _criar_xlsx(xlsx, ["uf", "", "sindicato"], [["SP", "x", "Sind"]])
    with pytest.raises(ErroCabecalhoPricing, match="branco"):
        ler_base_pricing(xlsx)


def test_reader_cabecalho_duplicado_levanta_erro(tmp_path):
    xlsx = tmp_path / "pricing.xlsx"
    _criar_xlsx(xlsx, ["uf", "sindicato", "uf"], [["SP", "Sind", "RJ"]])
    with pytest.raises(ErroCabecalhoPricing, match="duplicado"):
        ler_base_pricing(xlsx)


# ── integração via CLI ────────────────────────────────────────────────────────

def test_cli_fluxo_completo_cria_xlsx(tmp_path):
    from src.cli import main
    pricing = tmp_path / "pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato", "ano_referencia"],
                [["SP", "SindTest", "2024"], ["RJ", "Outro", "2023"]])
    adj = tmp_path / "adj.json"
    _criar_json_aprovados(adj, [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")])
    output = tmp_path / "out.xlsx"

    codigo = main([
        "preview-pricing-update",
        "--adjustments", str(adj),
        "--pricing", str(pricing),
        "--output", str(output),
    ])

    assert codigo == 0
    assert output.exists()


def test_cli_output_xlsx_tem_linhas_corretas(tmp_path):
    from src.cli import main
    pricing = tmp_path / "pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato", "ano_referencia"],
                [["SP", "SindTest", "2024"], ["RJ", "Outro", "2023"]])
    adj = tmp_path / "adj.json"
    _criar_json_aprovados(adj, [_aprovado(uf="SP", sindicato="SindTest", ano_referencia="2024")])
    output = tmp_path / "out.xlsx"

    main([
        "preview-pricing-update",
        "--adjustments", str(adj),
        "--pricing", str(pricing),
        "--output", str(output),
    ])

    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Linha 1 = cabeçalho, linhas 2 e 3 = dados
    assert ws.max_row == 3


def test_cli_relatorio_exibido_no_stdout(tmp_path, capsys):
    from src.cli import main
    pricing = tmp_path / "pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato"], [["SP", "SindTest"]])
    adj = tmp_path / "adj.json"
    _criar_json_aprovados(adj, [_aprovado()])
    output = tmp_path / "out.xlsx"

    main([
        "preview-pricing-update",
        "--adjustments", str(adj),
        "--pricing", str(pricing),
        "--output", str(output),
    ])
    captured = capsys.readouterr()
    assert "Relatório" in captured.out or "relat" in captured.out.lower()


def test_cli_caminhos_padrao_resolvidos_via_raiz_repo(tmp_path, monkeypatch):
    from src.cli import main
    import src.cli as cli_module

    monkeypatch.setattr(cli_module, "_raiz_repo", lambda: tmp_path)

    data_dir = tmp_path / "data"
    data_dir.mkdir()

    pricing = data_dir / "base_pricing.xlsx"
    _criar_xlsx(pricing, ["uf", "sindicato"], [["SP", "SindTest"]])
    _criar_json_aprovados(data_dir / "reajustes_aprovados.json", [_aprovado()])

    codigo = main(["preview-pricing-update"])
    assert codigo == 0
    assert (data_dir / "preview_atualizacao_pricing.xlsx").exists()
