"""
Testes automatizados para apply_review_decisions.py (PRJ-72).

Cobre os 14 cenários obrigatórios definidos nos ACs:
  1.  Dry-run não altera base_parametros_sindicais.json, .js nem audit.json
  2.  Execução real altera apenas campos do Excel com decisao_final válida
  3.  Erro de coluna obrigatória ausente aborta sem alterar nenhum arquivo
  4.  Decisão inválida gera erro na auditoria; demais linhas processadas
  5.  Campo inexistente gera erro na auditoria; demais linhas processadas
  6.  registro_id inexistente gera erro na auditoria; demais linhas processadas
  7.  validar com valor_revisado vazio atualiza apenas metadados, não altera valor
  8.  validar com valor_revisado diferente preserva valor_original_pre_validacao e atualiza valor
  9.  manter_pendente não valida (status permanece "pendente_revisao")
  10. rejeitar marca como "rejeitado"
  11. marcar_conflito preserva opcoes_identificadas existentes
  12. buscar_fonte mantém "pendente_revisao" com acao_recomendada: "buscar_fonte"
  13. Auditoria gerada contém todos os campos obrigatórios (antes/depois, revisor, timestamp)
  14. Base JS regenerada corretamente após execução real
"""

import copy
import json
import os
import sys
import tempfile
import unittest
from io import StringIO
from unittest.mock import patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from apply_review_decisions import (
    _apply_single_decision,
    _build_summary,
    _coerce_valor,
    apply_decisions,
    load_base,
    main,
    read_xlsx_decisions,
    regenerate_js,
    save_audit,
    save_base_json,
)

# ──────────────────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────────────────

TIMESTAMP_FIXO = "2026-06-18T12:00:00Z"
REVISOR = "Ana Silva"
DATA_REVISAO = "2026-06-17"


def _make_item(
    status: str = "pendente_revisao",
    valor=1000.0,
    opcoes=None,
    acao_recomendada=None,
) -> dict:
    item = {
        "valor": valor,
        "percentual": None,
        "valor_textual": None,
        "status_parametro": status,
        "origem": "pdf_cct",
        "fonte": "PDF da CCT",
    }
    if opcoes is not None:
        item["opcoes_identificadas"] = opcoes
    if acao_recomendada is not None:
        item["acao_recomendada"] = acao_recomendada
    return item


def _make_base(extra_registros: list | None = None) -> dict:
    registros = [
        {
            "id_registro_reajuste": "REG-SP-TEST-2025",
            "uf": "SP",
            "sindicato": "Sindicato Teste SP",
            "status_parametro": "pendente_revisao",
            "itens_cct": {
                "piso_salarial": _make_item(
                    status="pendente_revisao", valor=2000.0
                ),
                "adicional_noturno": _make_item(
                    status="pendente_revisao", valor=None
                ),
                "plr": _make_item(
                    status="pendente_revisao",
                    valor=500.0,
                    opcoes=[{"opcao": "A", "valor": 500.0}, {"opcao": "B", "valor": 600.0}],
                ),
                "hora_extra": _make_item(status="pendente_revisao", valor=50.0),
            },
        }
    ]
    if extra_registros:
        registros.extend(extra_registros)
    return {"data_geracao": "2026-06-18", "registros": registros}


def _make_decision(
    registro_id: str = "REG-SP-TEST-2025",
    campo: str = "piso_salarial",
    decisao_final: str = "validar",
    valor_revisado=None,
    observacao_revisor: str = "",
    revisor: str = REVISOR,
    data_revisao=DATA_REVISAO,
) -> dict:
    return {
        "registro_id": registro_id,
        "campo": campo,
        "decisao_final": decisao_final,
        "valor_revisado": valor_revisado,
        "observacao_revisor": observacao_revisor,
        "revisor": revisor,
        "data_revisao": data_revisao,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 1 — Dry-run não altera nenhum arquivo
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario01DryRunSemEfeitosColaterais(unittest.TestCase):
    def test_dry_run_nao_altera_arquivos(self):
        """Com --dry-run, nenhum arquivo deve ser criado ou modificado (AC3)."""
        decisions = [_make_decision()]

        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = os.path.join(tmpdir, "data")
            reports_dir = os.path.join(tmpdir, "reports")
            os.makedirs(data_dir)
            os.makedirs(reports_dir)

            json_path = os.path.join(data_dir, "base_parametros_sindicais.json")
            js_path = os.path.join(data_dir, "base_parametros_sindicais.js")
            audit_path = os.path.join(reports_dir, "review_decisions_audit.json")

            original_base = _make_base()
            save_base_json(original_base, json_path)

            original_json_content = open(json_path).read()

            # Simula dry-run: read_xlsx retorna decisions; arquivos são os do tmpdir
            with patch("apply_review_decisions.BASE_JSON_PATH", json_path), \
                 patch("apply_review_decisions.BASE_JS_PATH", js_path), \
                 patch("apply_review_decisions.AUDIT_PATH", audit_path), \
                 patch("apply_review_decisions.read_xlsx_decisions", return_value=decisions):

                ret = main(["--decisions", "fake.xlsx", "--dry-run"])

            self.assertEqual(ret, 0)
            # JSON não modificado
            self.assertEqual(open(json_path).read(), original_json_content)
            # JS não criado
            self.assertFalse(os.path.exists(js_path))
            # Auditoria não criada
            self.assertFalse(os.path.exists(audit_path))


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 2 — Execução real altera apenas campos com decisao_final válida
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario02ExecucaoRealAlteraCamposCertos(unittest.TestCase):
    def test_apenas_campos_decididos_sao_alterados(self):
        """Apenas campos presentes no Excel com decisao_final válida são alterados (AC5)."""
        base = _make_base()
        decisions = [_make_decision(campo="piso_salarial", decisao_final="validar")]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        registro = base_modificada["registros"][0]
        # Campo decidido: status atualizado
        self.assertEqual(
            registro["itens_cct"]["piso_salarial"]["status_parametro"], "valido"
        )
        # Campos não decididos: status inalterado
        self.assertEqual(
            registro["itens_cct"]["adicional_noturno"]["status_parametro"],
            "pendente_revisao",
        )
        self.assertEqual(
            registro["itens_cct"]["plr"]["status_parametro"], "pendente_revisao"
        )
        self.assertEqual(len(audit), 1)
        self.assertEqual(audit[0]["resultado"], "aplicado")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 3 — Coluna obrigatória ausente aborta sem alterar arquivos
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario03ColunaAusenteAbortaExecucao(unittest.TestCase):
    def test_coluna_decisao_final_ausente_aborta(self):
        """Coluna obrigatória ausente → script retorna código não-zero sem alterar arquivos (AC1)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = os.path.join(tmpdir, "data")
            reports_dir = os.path.join(tmpdir, "reports")
            os.makedirs(data_dir)
            os.makedirs(reports_dir)

            json_path = os.path.join(data_dir, "base_parametros_sindicais.json")
            js_path = os.path.join(data_dir, "base_parametros_sindicais.js")
            audit_path = os.path.join(reports_dir, "review_decisions_audit.json")

            save_base_json(_make_base(), json_path)
            original_content = open(json_path).read()

            # read_xlsx_decisions levanta ValueError (coluna ausente)
            def mock_read_xlsx_missing_col(path):
                raise ValueError("Colunas obrigatórias ausentes no Excel: decisao_final")

            with patch("apply_review_decisions.BASE_JSON_PATH", json_path), \
                 patch("apply_review_decisions.BASE_JS_PATH", js_path), \
                 patch("apply_review_decisions.AUDIT_PATH", audit_path), \
                 patch(
                     "apply_review_decisions.read_xlsx_decisions",
                     side_effect=mock_read_xlsx_missing_col,
                 ):
                ret = main(["--decisions", "fake.xlsx"])

            self.assertNotEqual(ret, 0)
            self.assertEqual(open(json_path).read(), original_content)
            self.assertFalse(os.path.exists(js_path))
            self.assertFalse(os.path.exists(audit_path))

    def test_read_xlsx_levanta_erro_quando_coluna_ausente(self):
        """read_xlsx_decisions levanta ValueError descritivo se coluna de revisão estiver ausente."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl não disponível")

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = os.path.join(tmpdir, "missing_col.xlsx")

            wb = openpyxl.Workbook()
            ws = wb.active
            # Colunas de revisão incompletas (falta decisao_final e outras)
            ws.append(["registro_id", "campo", "valor_revisado"])
            ws.append(["REG-SP-TEST-2025", "piso_salarial", "2000.0"])
            wb.save(xlsx_path)

            with self.assertRaises(ValueError) as ctx:
                read_xlsx_decisions(xlsx_path)

            self.assertIn("ausentes", str(ctx.exception).lower())


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 4 — Decisão inválida gera erro; demais linhas processadas
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario04DecisaoInvalidaIsolaErro(unittest.TestCase):
    def test_decisao_invalida_nao_altera_campo_e_demais_sao_processadas(self):
        """Decisão inválida → erro na auditoria; campo não alterado; outras linhas continuam (AC2)."""
        base = _make_base()
        decisions = [
            _make_decision(campo="piso_salarial", decisao_final="aprovado"),  # inválida
            _make_decision(campo="hora_extra", decisao_final="rejeitar"),  # válida
        ]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        self.assertEqual(len(audit), 2)
        self.assertEqual(audit[0]["resultado"], "erro")
        self.assertIn("inválida", audit[0]["motivo"])
        # Campo com decisão inválida: não alterado
        self.assertEqual(
            base_modificada["registros"][0]["itens_cct"]["piso_salarial"]["status_parametro"],
            "pendente_revisao",
        )
        # Campo válido: alterado
        self.assertEqual(audit[1]["resultado"], "aplicado")
        self.assertEqual(
            base_modificada["registros"][0]["itens_cct"]["hora_extra"]["status_parametro"],
            "rejeitado",
        )


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 5 — Campo inexistente gera erro; demais linhas processadas
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario05CampoInexistenteIsolaErro(unittest.TestCase):
    def test_campo_inexistente_gera_erro_demais_processadas(self):
        """Campo não encontrado em itens_cct → erro de auditoria; outras linhas continuam (AC2)."""
        base = _make_base()
        decisions = [
            _make_decision(campo="campo_fantasma", decisao_final="validar"),  # inexistente
            _make_decision(campo="piso_salarial", decisao_final="validar"),  # válido
        ]

        _, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        self.assertEqual(audit[0]["resultado"], "erro")
        self.assertIn("campo_fantasma", audit[0]["motivo"])
        self.assertEqual(audit[1]["resultado"], "aplicado")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 6 — registro_id inexistente gera erro; demais linhas processadas
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario06RegistroInexistenteIsolaErro(unittest.TestCase):
    def test_registro_id_inexistente_gera_erro_demais_processadas(self):
        """registro_id não encontrado → erro de auditoria; outras linhas continuam (AC2)."""
        base = _make_base()
        decisions = [
            _make_decision(
                registro_id="REG-FANTASMA-2025", campo="piso_salarial", decisao_final="validar"
            ),  # inexistente
            _make_decision(campo="hora_extra", decisao_final="manter_pendente"),  # válido
        ]

        _, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        self.assertEqual(audit[0]["resultado"], "erro")
        self.assertIn("REG-FANTASMA-2025", audit[0]["motivo"])
        self.assertEqual(audit[1]["resultado"], "aplicado")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 7 — validar com valor_revisado vazio: apenas metadados atualizados
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario07ValidarSemValorRevisado(unittest.TestCase):
    def test_validar_sem_valor_revisado_nao_altera_valor(self):
        """validar + valor_revisado vazio → apenas metadados de status; valor não tocado (AC2)."""
        base = _make_base()
        valor_original = base["registros"][0]["itens_cct"]["piso_salarial"]["valor"]

        decisions = [
            _make_decision(
                campo="piso_salarial",
                decisao_final="validar",
                valor_revisado=None,
                revisor=REVISOR,
            )
        ]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["status_parametro"], "valido")
        self.assertEqual(item["valor"], valor_original)
        self.assertNotIn("valor_original_pre_validacao", item)
        self.assertEqual(item["validado_por"], REVISOR)
        self.assertEqual(audit[0]["resultado"], "aplicado")

    def test_validar_com_valor_revisado_igual_ao_atual_nao_altera_valor(self):
        """validar + valor_revisado igual ao valor atual → apenas metadados (AC2)."""
        base = _make_base()
        valor_atual = base["registros"][0]["itens_cct"]["piso_salarial"]["valor"]  # 2000.0

        decisions = [
            _make_decision(
                campo="piso_salarial",
                decisao_final="validar",
                valor_revisado=valor_atual,  # igual
            )
        ]

        base_modificada, _ = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["status_parametro"], "valido")
        self.assertEqual(item["valor"], valor_atual)
        self.assertNotIn("valor_original_pre_validacao", item)

    def test_validar_com_valor_revisado_string_vazia(self):
        """validar + valor_revisado = '' → apenas metadados (AC2)."""
        base = _make_base()
        valor_original = base["registros"][0]["itens_cct"]["piso_salarial"]["valor"]

        decisions = [
            _make_decision(
                campo="piso_salarial",
                decisao_final="validar",
                valor_revisado="",
            )
        ]

        base_modificada, _ = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["valor"], valor_original)
        self.assertNotIn("valor_original_pre_validacao", item)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 8 — validar com valor_revisado diferente preserva valor_original e atualiza valor
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario08ValidarComValorRevisadoDiferente(unittest.TestCase):
    def test_valor_revisado_diferente_preserva_original_e_atualiza(self):
        """validar + valor_revisado diferente → copia valor_original e atualiza valor (AC2)."""
        base = _make_base()
        valor_anterior = base["registros"][0]["itens_cct"]["piso_salarial"]["valor"]  # 2000.0
        novo_valor = 2200.0

        decisions = [
            _make_decision(
                campo="piso_salarial",
                decisao_final="validar",
                valor_revisado=novo_valor,
            )
        ]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["status_parametro"], "valido")
        self.assertEqual(item["valor"], novo_valor)
        self.assertEqual(item["valor_original_pre_validacao"], valor_anterior)
        self.assertEqual(audit[0]["resultado"], "aplicado")
        self.assertEqual(audit[0]["valor_anterior"], valor_anterior)
        self.assertEqual(audit[0]["valor_novo"], novo_valor)

    def test_valor_revisado_como_string_numeric_eh_coercido(self):
        """valor_revisado string numérico é convertido para float antes da comparação (AC2)."""
        base = _make_base()
        decisions = [
            _make_decision(
                campo="piso_salarial",
                decisao_final="validar",
                valor_revisado="2200.00",
            )
        ]

        base_modificada, _ = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["valor"], 2200.0)
        self.assertEqual(item["valor_original_pre_validacao"], 2000.0)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 9 — manter_pendente não valida
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario09ManterPendente(unittest.TestCase):
    def test_manter_pendente_nao_valida(self):
        """manter_pendente → status_parametro permanece 'pendente_revisao' (AC2)."""
        base = _make_base()
        decisions = [_make_decision(campo="piso_salarial", decisao_final="manter_pendente")]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["status_parametro"], "pendente_revisao")
        self.assertNotIn("validado_por", item)
        self.assertEqual(audit[0]["resultado"], "aplicado")
        self.assertEqual(audit[0]["status_novo"], "pendente_revisao")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 10 — rejeitar marca como "rejeitado"
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario10Rejeitar(unittest.TestCase):
    def test_rejeitar_marca_status_rejeitado(self):
        """rejeitar → status_parametro = 'rejeitado' (AC2)."""
        base = _make_base()
        decisions = [_make_decision(campo="piso_salarial", decisao_final="rejeitar")]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["piso_salarial"]
        self.assertEqual(item["status_parametro"], "rejeitado")
        self.assertEqual(audit[0]["resultado"], "aplicado")
        self.assertEqual(audit[0]["status_novo"], "rejeitado")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 11 — marcar_conflito preserva opcoes_identificadas
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario11MarcarConflito(unittest.TestCase):
    def test_marcar_conflito_preserva_opcoes_identificadas(self):
        """marcar_conflito → status = 'conflito'; opcoes_identificadas intactas (AC2)."""
        base = _make_base()
        opcoes_originais = base["registros"][0]["itens_cct"]["plr"]["opcoes_identificadas"]
        self.assertIsNotNone(opcoes_originais)

        decisions = [_make_decision(campo="plr", decisao_final="marcar_conflito")]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["plr"]
        self.assertEqual(item["status_parametro"], "conflito")
        self.assertEqual(item["opcoes_identificadas"], opcoes_originais)
        self.assertEqual(audit[0]["resultado"], "aplicado")

    def test_marcar_conflito_sem_opcoes_nao_cria_opcoes(self):
        """marcar_conflito sem opcoes_identificadas → não adiciona a chave (AC2)."""
        base = _make_base()
        decisions = [_make_decision(campo="hora_extra", decisao_final="marcar_conflito")]

        base_modificada, _ = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["hora_extra"]
        self.assertEqual(item["status_parametro"], "conflito")
        self.assertNotIn("opcoes_identificadas", item)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 12 — buscar_fonte mantém "pendente_revisao" com acao_recomendada
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario12BuscarFonte(unittest.TestCase):
    def test_buscar_fonte_status_e_acao_recomendada(self):
        """buscar_fonte → status 'pendente_revisao' + acao_recomendada: 'buscar_fonte' (AC2)."""
        base = _make_base()
        decisions = [_make_decision(campo="adicional_noturno", decisao_final="buscar_fonte")]

        base_modificada, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        item = base_modificada["registros"][0]["itens_cct"]["adicional_noturno"]
        self.assertEqual(item["status_parametro"], "pendente_revisao")
        self.assertEqual(item.get("acao_recomendada"), "buscar_fonte")
        self.assertEqual(audit[0]["resultado"], "aplicado")
        self.assertEqual(audit[0]["status_novo"], "pendente_revisao")


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 13 — Auditoria contém todos os campos obrigatórios
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario13AuditoriaComCamposObrigatorios(unittest.TestCase):
    CAMPOS_OBRIGATORIOS = {
        "registro_id",
        "campo",
        "decisao_final",
        "status_anterior",
        "status_novo",
        "valor_anterior",
        "valor_novo",
        "revisor",
        "data_revisao",
        "observacao_revisor",
        "resultado",
        "motivo",
        "timestamp_execucao",
    }

    def test_registro_aplicado_tem_todos_os_campos(self):
        """Registro de auditoria 'aplicado' contém todos os campos obrigatórios (AC4)."""
        base = _make_base()
        decisions = [
            _make_decision(
                campo="piso_salarial",
                decisao_final="validar",
                observacao_revisor="Verificado",
            )
        ]

        _, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        self.assertEqual(len(audit), 1)
        self.assertSetEqual(
            self.CAMPOS_OBRIGATORIOS, self.CAMPOS_OBRIGATORIOS & set(audit[0].keys())
        )
        self.assertEqual(audit[0]["timestamp_execucao"], TIMESTAMP_FIXO)
        self.assertEqual(audit[0]["revisor"], REVISOR)
        self.assertEqual(audit[0]["status_anterior"], "pendente_revisao")
        self.assertEqual(audit[0]["status_novo"], "valido")

    def test_registro_erro_tem_todos_os_campos(self):
        """Registro de auditoria 'erro' também contém todos os campos obrigatórios (AC4)."""
        base = _make_base()
        decisions = [_make_decision(campo="piso_salarial", decisao_final="INVALIDA")]

        _, audit = apply_decisions(
            copy.deepcopy(base), decisions, timestamp=TIMESTAMP_FIXO
        )

        self.assertEqual(audit[0]["resultado"], "erro")
        self.assertSetEqual(
            self.CAMPOS_OBRIGATORIOS, self.CAMPOS_OBRIGATORIOS & set(audit[0].keys())
        )
        self.assertIsNotNone(audit[0]["motivo"])

    def test_auditoria_persistida_em_arquivo(self):
        """save_audit grava JSON válido com todos os registros."""
        with tempfile.TemporaryDirectory() as tmpdir:
            audit_path = os.path.join(tmpdir, "audit.json")
            records = [
                {
                    "registro_id": "REG-SP-TEST-2025",
                    "campo": "piso_salarial",
                    "decisao_final": "validar",
                    "status_anterior": "pendente_revisao",
                    "status_novo": "valido",
                    "valor_anterior": 2000.0,
                    "valor_novo": 2000.0,
                    "revisor": REVISOR,
                    "data_revisao": DATA_REVISAO,
                    "observacao_revisor": "",
                    "resultado": "aplicado",
                    "motivo": None,
                    "timestamp_execucao": TIMESTAMP_FIXO,
                }
            ]
            save_audit(records, audit_path)
            self.assertTrue(os.path.exists(audit_path))
            loaded = json.load(open(audit_path))
            self.assertEqual(loaded, records)


# ──────────────────────────────────────────────────────────────────────────────
# Cenário 14 — Base JS regenerada corretamente após execução real
# ──────────────────────────────────────────────────────────────────────────────


class TestCenario14JSRegenerado(unittest.TestCase):
    def test_js_regenerado_apos_execucao_real(self):
        """Após execução real, base_parametros_sindicais.js contém os dados atualizados (AC5)."""
        decisions = [_make_decision(campo="piso_salarial", decisao_final="validar")]

        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = os.path.join(tmpdir, "data")
            reports_dir = os.path.join(tmpdir, "reports")
            os.makedirs(data_dir)
            os.makedirs(reports_dir)

            json_path = os.path.join(data_dir, "base_parametros_sindicais.json")
            js_path = os.path.join(data_dir, "base_parametros_sindicais.js")
            audit_path = os.path.join(reports_dir, "review_decisions_audit.json")

            original_base = _make_base()
            save_base_json(original_base, json_path)

            with patch("apply_review_decisions.BASE_JSON_PATH", json_path), \
                 patch("apply_review_decisions.BASE_JS_PATH", js_path), \
                 patch("apply_review_decisions.AUDIT_PATH", audit_path), \
                 patch("apply_review_decisions.read_xlsx_decisions", return_value=decisions):

                ret = main(["--decisions", "fake.xlsx"])

            self.assertEqual(ret, 0)

            # JS criado
            self.assertTrue(os.path.exists(js_path))
            js_content = open(js_path).read()
            self.assertTrue(js_content.startswith("// Gerado automaticamente"))
            self.assertIn("window.BASE_PARAMETROS_SINDICAIS = ", js_content)

            # JS contém dado atualizado (status valido)
            self.assertIn('"valido"', js_content)

            # JSON atualizado
            base_salva = json.load(open(json_path))
            status = base_salva["registros"][0]["itens_cct"]["piso_salarial"]["status_parametro"]
            self.assertEqual(status, "valido")

            # Auditoria criada
            self.assertTrue(os.path.exists(audit_path))

    def test_regenerate_js_gera_conteudo_correto(self):
        """regenerate_js gera conteúdo idêntico ao de export_inline_data.py."""
        data = _make_base()
        with tempfile.TemporaryDirectory() as tmpdir:
            js_path = os.path.join(tmpdir, "test.js")
            regenerate_js(data, js_path)
            content = open(js_path).read()

        expected_prefix = "// Gerado automaticamente por export_inline_data.py — não editar manualmente.\n"
        expected_var = "window.BASE_PARAMETROS_SINDICAIS = "
        self.assertTrue(content.startswith(expected_prefix))
        self.assertIn(expected_var, content)
        self.assertTrue(content.endswith(";\n"))
        # Dados embutidos são válidos JSON
        json_part = content[content.index(expected_var) + len(expected_var):-2]
        parsed = json.loads(json_part)
        self.assertEqual(parsed["registros"][0]["id_registro_reajuste"], "REG-SP-TEST-2025")


# ──────────────────────────────────────────────────────────────────────────────
# Testes adicionais de leitura XLSX (requerem openpyxl)
# ──────────────────────────────────────────────────────────────────────────────


class TestLeituraXLSXPerCampo(unittest.TestCase):
    """Testa read_xlsx_decisions com XLSX no formato por-campo."""

    def _make_xlsx(self, tmpdir: str, rows: list[list]) -> str:
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl não disponível")

        path = os.path.join(tmpdir, "test.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            ["registro_id", "campo", "decisao_final", "valor_revisado",
             "observacao_revisor", "revisor", "data_revisao"]
        )
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    def test_leitura_formato_por_campo(self):
        """Lê XLSX formato por-campo retornando dicts com campos corretos."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = self._make_xlsx(
                tmpdir,
                [["REG-SP-TEST-2025", "piso_salarial", "validar", 2200.0, "Ok", "Ana", "2026-06-17"]],
            )
            decisions = read_xlsx_decisions(xlsx_path)

        self.assertEqual(len(decisions), 1)
        self.assertEqual(decisions[0]["registro_id"], "REG-SP-TEST-2025")
        self.assertEqual(decisions[0]["campo"], "piso_salarial")
        self.assertEqual(decisions[0]["decisao_final"], "validar")
        self.assertEqual(decisions[0]["valor_revisado"], 2200.0)
        self.assertEqual(decisions[0]["revisor"], "Ana")

    def test_linhas_completamente_vazias_ignoradas(self):
        """Linhas completamente vazias no XLSX são ignoradas."""
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path = self._make_xlsx(
                tmpdir,
                [
                    ["REG-SP-TEST-2025", "piso_salarial", "validar", None, "", "Ana", None],
                    [None, None, None, None, None, None, None],  # linha vazia
                    ["REG-SP-TEST-2025", "hora_extra", "rejeitar", None, "", "Ana", None],
                ],
            )
            decisions = read_xlsx_decisions(xlsx_path)

        self.assertEqual(len(decisions), 2)

    def test_colunas_ausentes_levantam_value_error(self):
        """Se colunas obrigatórias de revisão estiverem ausentes, ValueError é levantado (AC1)."""
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl não disponível")

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "incomplete.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["registro_id", "campo", "decisao_final"])  # faltam colunas
            wb.save(path)

            with self.assertRaises(ValueError) as ctx:
                read_xlsx_decisions(path)

            self.assertIn("ausentes", str(ctx.exception).lower())

    def test_arquivo_inexistente_levanta_file_not_found(self):
        """Arquivo inexistente levanta FileNotFoundError."""
        with self.assertRaises(FileNotFoundError):
            read_xlsx_decisions("/tmp/nao_existe_XYZ123.xlsx")


# ──────────────────────────────────────────────────────────────────────────────
# Testes auxiliares: _coerce_valor e _build_summary
# ──────────────────────────────────────────────────────────────────────────────


class TestCoerceValor(unittest.TestCase):
    def test_float_passado_diretamente(self):
        self.assertEqual(_coerce_valor(1500.5, 1000.0), 1500.5)

    def test_string_decimal_ponto(self):
        self.assertEqual(_coerce_valor("1500.50", 1000.0), 1500.50)

    def test_string_decimal_virgula(self):
        self.assertAlmostEqual(_coerce_valor("1.500,47", 1000.0), 1500.47, places=2)

    def test_string_nao_numerica_retorna_original(self):
        self.assertEqual(_coerce_valor("abc", 1000.0), "abc")

    def test_int_anterior_retorna_int(self):
        result = _coerce_valor(44, 40)
        self.assertEqual(result, 44)


class TestBuildSummary(unittest.TestCase):
    def test_contagens_corretas(self):
        records = [
            {"resultado": "aplicado", "decisao_final": "validar"},
            {"resultado": "aplicado", "decisao_final": "validar"},
            {"resultado": "aplicado", "decisao_final": "rejeitar"},
            {"resultado": "aplicado", "decisao_final": "manter_pendente"},
            {"resultado": "aplicado", "decisao_final": "buscar_fonte"},
            {"resultado": "aplicado", "decisao_final": "marcar_conflito"},
            {"resultado": "erro", "decisao_final": "invalida"},
        ]
        summary = _build_summary(records)
        self.assertEqual(summary["total_lidas"], 7)
        self.assertEqual(summary["validar"], 2)
        self.assertEqual(summary["rejeitar"], 1)
        self.assertEqual(summary["manter_pendente"], 1)
        self.assertEqual(summary["buscar_fonte"], 1)
        self.assertEqual(summary["marcar_conflito"], 1)
        self.assertEqual(summary["erro"], 1)


# ──────────────────────────────────────────────────────────────────────────────
# Testes de integração: execução real completa via main()
# ──────────────────────────────────────────────────────────────────────────────


class TestIntegracaoExecucaoReal(unittest.TestCase):
    def _run_main_with_decisions(self, decisions: list[dict], dry_run: bool = False):
        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = os.path.join(tmpdir, "data")
            reports_dir = os.path.join(tmpdir, "reports")
            os.makedirs(data_dir)
            os.makedirs(reports_dir)

            json_path = os.path.join(data_dir, "base_parametros_sindicais.json")
            js_path = os.path.join(data_dir, "base_parametros_sindicais.js")
            audit_path = os.path.join(reports_dir, "review_decisions_audit.json")

            save_base_json(_make_base(), json_path)

            argv = ["--decisions", "fake.xlsx"]
            if dry_run:
                argv.append("--dry-run")

            with patch("apply_review_decisions.BASE_JSON_PATH", json_path), \
                 patch("apply_review_decisions.BASE_JS_PATH", js_path), \
                 patch("apply_review_decisions.AUDIT_PATH", audit_path), \
                 patch("apply_review_decisions.read_xlsx_decisions", return_value=decisions):
                ret = main(argv)

            return ret, json_path, js_path, audit_path

    def test_execucao_real_retorna_zero(self):
        decisions = [_make_decision(campo="piso_salarial", decisao_final="validar")]
        ret, *_ = self._run_main_with_decisions(decisions)
        self.assertEqual(ret, 0)

    def test_multiplas_decisoes_aplicadas_em_sequencia(self):
        """Múltiplas decisões válidas são todas aplicadas (AC2)."""
        decisions = [
            _make_decision(campo="piso_salarial", decisao_final="validar"),
            _make_decision(campo="hora_extra", decisao_final="rejeitar"),
            _make_decision(campo="plr", decisao_final="marcar_conflito"),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            data_dir = os.path.join(tmpdir, "data")
            reports_dir = os.path.join(tmpdir, "reports")
            os.makedirs(data_dir)
            os.makedirs(reports_dir)

            json_path = os.path.join(data_dir, "base_parametros_sindicais.json")
            js_path = os.path.join(data_dir, "base_parametros_sindicais.js")
            audit_path = os.path.join(reports_dir, "review_decisions_audit.json")

            save_base_json(_make_base(), json_path)

            with patch("apply_review_decisions.BASE_JSON_PATH", json_path), \
                 patch("apply_review_decisions.BASE_JS_PATH", js_path), \
                 patch("apply_review_decisions.AUDIT_PATH", audit_path), \
                 patch("apply_review_decisions.read_xlsx_decisions", return_value=decisions):
                ret = main(["--decisions", "fake.xlsx"])

            self.assertEqual(ret, 0)
            base = json.load(open(json_path))
            itens = base["registros"][0]["itens_cct"]
            self.assertEqual(itens["piso_salarial"]["status_parametro"], "valido")
            self.assertEqual(itens["hora_extra"]["status_parametro"], "rejeitado")
            self.assertEqual(itens["plr"]["status_parametro"], "conflito")

            audit = json.load(open(audit_path))
            self.assertEqual(len(audit), 3)
            self.assertTrue(all(r["resultado"] == "aplicado" for r in audit))

    def test_dry_run_nao_cria_arquivos(self):
        """Dry-run não cria JS nem audit (AC3)."""
        decisions = [_make_decision(campo="piso_salarial", decisao_final="validar")]
        ret, json_path, js_path, audit_path = self._run_main_with_decisions(
            decisions, dry_run=True
        )
        self.assertEqual(ret, 0)
        self.assertFalse(os.path.exists(js_path))
        self.assertFalse(os.path.exists(audit_path))

    def test_base_inexistente_retorna_codigo_nao_zero(self):
        """Base inexistente → código de saída não-zero."""
        with patch(
            "apply_review_decisions.BASE_JSON_PATH", "/tmp/nao_existe_XYZ.json"
        ), patch(
            "apply_review_decisions.read_xlsx_decisions",
            return_value=[_make_decision()],
        ):
            ret = main(["--decisions", "fake.xlsx"])
        self.assertNotEqual(ret, 0)


if __name__ == "__main__":
    unittest.main()
