"""Leitura, revisão e filtragem da prévia de atualização de pricing.

Responsabilidades:
  - Ler o arquivo ``data/preview_atualizacao_pricing.xlsx`` como lista de dicts.
  - Garantir a existência da coluna ``decisao_aplicacao`` (adicionando-a vazia
    quando ausente) — AC1-a.
  - Filtrar exclusivamente as linhas com ``decisao_aplicacao = aprovado`` **e**
    ``status_aplicacao = reajuste_encontrado`` — AC2, AC6.
  - Gravar atomicamente a base de aplicações aprovadas — AC2, AC5.
"""

import os
import tempfile
from pathlib import Path
from typing import List, Tuple

import openpyxl

COLUNA_DECISAO = "decisao_aplicacao"
COLUNA_STATUS = "status_aplicacao"
STATUS_ELEGIVEL = "reajuste_encontrado"
VALOR_APROVADO = "aprovado"


def ler_preview_xlsx(path: Path) -> Tuple[List[dict], List[str]]:
    """Lê todas as linhas do arquivo de prévia de pricing.

    Returns:
        linhas  — lista de dicts ``{coluna: valor}`` por linha de dados.
        colunas — lista ordenada de nomes de coluna conforme o cabeçalho.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    colunas = [str(c) if c is not None else "" for c in rows[0]]
    linhas = [{colunas[i]: row[i] for i in range(len(colunas))} for row in rows[1:]]
    return linhas, colunas


def garantir_coluna_decisao_aplicacao(path: Path) -> bool:
    """Garante que ``decisao_aplicacao`` exista no arquivo de prévia.

    Se a coluna já existir, não altera o arquivo e retorna False.
    Se estiver ausente, adiciona a coluna com células em branco em todas as
    linhas de dados, persiste o arquivo de forma atômica e retorna True.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    cabecalho = [str(ws.cell(1, c).value) if ws.cell(1, c).value is not None else ""
                 for c in range(1, ws.max_column + 1)]

    if COLUNA_DECISAO in cabecalho:
        wb.close()
        return False

    nova_col = ws.max_column + 1
    ws.cell(row=1, column=nova_col, value=COLUNA_DECISAO)
    # células de dados já são None (em branco) por padrão — nenhuma ação adicional

    _salvar_xlsx_atomico(wb, path)
    return True


def filtrar_aprovadas_elegiveis(
    linhas: List[dict],
) -> Tuple[List[dict], int]:
    """Filtra linhas com ``decisao_aplicacao = aprovado`` e ``status_aplicacao = reajuste_encontrado``.

    Returns:
        aprovadas          — linhas que satisfazem ambos os critérios.
        total_inelegiveis  — contagem de linhas com ``decisao_aplicacao = aprovado``
                             mas ``status_aplicacao`` diferente de ``reajuste_encontrado``.
    """
    aprovadas: List[dict] = []
    total_inelegiveis = 0

    for linha in linhas:
        decisao = (linha.get(COLUNA_DECISAO) or "").strip().lower()
        status = (linha.get(COLUNA_STATUS) or "").strip()

        if decisao != VALOR_APROVADO:
            continue

        if status == STATUS_ELEGIVEL:
            aprovadas.append(linha)
        else:
            total_inelegiveis += 1

    return aprovadas, total_inelegiveis


def salvar_base_aplicacoes(
    output_path: Path,
    linhas: List[dict],
    colunas: List[str],
) -> None:
    """Grava a base de aplicações aprovadas de forma atômica.

    Escreve todas as colunas (originais + enriquecimento) na ordem fornecida
    por ``colunas``.  Se ``linhas`` for vazia, o arquivo é criado apenas com
    o cabeçalho — AC2.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "aplicacoes_aprovadas"

    ws.append(list(colunas))
    for linha in linhas:
        ws.append([linha.get(col) for col in colunas])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _salvar_xlsx_atomico(wb, output_path)


def _salvar_xlsx_atomico(wb: openpyxl.Workbook, dest: Path) -> None:
    """Persiste ``wb`` em ``dest`` via tempfile + os.replace (escrita atômica)."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=dest.parent, suffix=".xlsx.tmp")
    try:
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, dest)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise
