"""Gravação da prévia de atualização de pricing em formato XLSX.

Usa escrita atômica via ``tempfile + os.replace`` para não corromper nem
substituir o arquivo original — AC2.
"""

import os
import tempfile
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font

from src.models.linha_preview_pricing import COLUNAS_PREVIEW, LinhaPreviewPricing


def salvar_preview(
    output_path: Path,
    headers: List[str],
    linhas: List[LinhaPreviewPricing],
) -> None:
    """Grava prévia em XLSX de forma atômica.

    Preserva todos os cabeçalhos originais da base de pricing e acrescenta
    as 8 colunas de prévia ao final (AC4). O arquivo original não é tocado
    durante a escrita (AC2).

    Args:
        output_path: caminho do arquivo de saída.
        headers: cabeçalhos originais da base de pricing.
        linhas: lista de linhas com dados originais e campos de prévia.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(dir=output_path.parent, suffix=".tmp.xlsx")
    try:
        os.close(fd)
        _escrever_xlsx(Path(tmp_path), headers, linhas)
        os.replace(tmp_path, output_path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def _escrever_xlsx(
    path: Path,
    headers: List[str],
    linhas: List[LinhaPreviewPricing],
) -> None:
    """Escreve arquivo XLSX com cabeçalhos originais + colunas de prévia."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Preview Pricing"

    cabecalho_completo = headers + COLUNAS_PREVIEW

    # Linha de cabeçalho com negrito
    for col_idx, nome in enumerate(cabecalho_completo, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = Font(bold=True)

    # Linhas de dados
    for row_idx, linha in enumerate(linhas, start=2):
        # Colunas originais
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=linha.dados_originais.get(header))

        # Colunas de prévia (na ordem de COLUNAS_PREVIEW)
        offset = len(headers) + 1
        preview_values = [
            linha.id_registro_reajuste,
            linha.percentual_reajuste_final,
            linha.data_base_final,
            linha.vigencia_inicio_final,
            linha.vigencia_fim_final,
            linha.fonte_documento,
            linha.status_aplicacao,
            linha.observacao_aplicacao,
        ]
        for col_idx, valor in enumerate(preview_values, start=offset):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    wb.save(path)
